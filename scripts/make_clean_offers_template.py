"""
One-shot: take the latest offers-import xlsx uploaded under instance/uploads
and produce a clean header-only copy (keeping header styling, freeze panes,
column widths, etc.) at instance/repricing/offers_import_blank.xlsx.

Run this once on the server. The full-export service will load that file as
its base each time.
"""
import os
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook


def find_latest_template(uploads_dir: str) -> str:
    """Pick the most recent offers-import-*.xlsx that has the right header.
    Excludes result_*.xlsx (those are already filled outputs).
    """
    candidates = []
    for name in os.listdir(uploads_dir):
        if not name.endswith(".xlsx"):
            continue
        if not name.lower().startswith(("offers-import", "macy_offers-import")):
            continue
        if name.lower().startswith("result_"):
            continue
        path = os.path.join(uploads_dir, name)
        candidates.append((os.path.getmtime(path), path))
    if not candidates:
        raise SystemExit("No offers-import template found in uploads/")
    candidates.sort(reverse=True)
    return candidates[0][1]


def main():
    uploads_dir = "/var/www/autoweb/AutoWeb/instance/uploads"
    out_dir = "/var/www/autoweb/AutoWeb/instance/repricing"
    out_path = os.path.join(out_dir, "offers_import_blank.xlsx")

    os.makedirs(out_dir, exist_ok=True)

    src = find_latest_template(uploads_dir)
    print(f"source: {src}")

    # Copy then clear data rows 2..end
    shutil.copyfile(src, out_path)

    wb = load_workbook(out_path)
    sheet_name = "offers-import" if "offers-import" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    print(f"sheet: {sheet_name}")
    print(f"orig rows: {ws.max_row}, cols: {ws.max_column}")

    # Delete every row except row 1 (header)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    print(f"after clear rows: {ws.max_row}, cols: {ws.max_column}")
    print(f"freeze_panes: {ws.freeze_panes}")

    wb.save(out_path)
    print(f"\nwritten: {out_path} ({os.path.getsize(out_path)} bytes)")


if __name__ == "__main__":
    main()
