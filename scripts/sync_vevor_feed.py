# -*- coding: utf-8 -*-
"""Vevor feed 同步（2026-07-23，选品用）：
下载 vevor-533.xlsx → 存 autooperate.vevor_feed（SKU/类目/库存/标题/图/价格/尺寸）。
Costway 的类目/库存已在 safety_product_cache + newestdropship；Vevor 独立存这里。
"""
import io
import os
import sys
from pathlib import Path

import requests
from openpyxl import load_workbook

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from app import create_app
from app.models.db_manager import DBManager

FEED_URL = "https://ads-feed.s3.us-west-2.amazonaws.com/ads/business/533/vevor-533.xlsx"

DDL = """
CREATE TABLE IF NOT EXISTS autooperate.vevor_feed (
    sku VARCHAR(64) PRIMARY KEY,
    product_type VARCHAR(400) COMMENT '供应商类目(层级)',
    title VARCHAR(400),
    inventory INT,
    price VARCHAR(32),
    image VARCHAR(600),
    long_in DECIMAL(8,2), wide_in DECIMAL(8,2), high_in DECIMAL(8,2),
    weight_lb DECIMAL(8,2),
    synced_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
    KEY idx_type (product_type(255)), KEY idx_inv (inventory)
) CHARSET=utf8mb4 COMMENT='Vevor全量feed(选品用)'
"""


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main(local_path=None) -> int:
    app = create_app(os.environ.get("FLASK_CONFIG", "production"))
    with app.app_context():
        conn = DBManager.get_connection()
        try:
            with conn.cursor() as cur:
                cur.execute(DDL)
            conn.commit()
        finally:
            conn.close()

        if local_path and os.path.exists(local_path):
            with open(local_path, "rb") as fh:
                content = fh.read()
        else:
            content = requests.get(FEED_URL, timeout=300).content
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb["feed"] if "feed" in wb.sheetnames else wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        head = list(next(it))
        idx = {h: i for i, h in enumerate(head)}

        def g(r, name):
            i = idx.get(name)
            return r[i] if i is not None and i < len(r) else None

        rows = []
        for r in it:
            sku = (g(r, "SKU") or "")
            sku = str(sku).strip() if sku else ""
            if not sku:
                continue
            inv = _f(g(r, "Inventory quantity"))
            rows.append((
                sku[:64], (g(r, "Product type") or "")[:400],
                (g(r, "Product title") or "")[:400],
                int(inv) if inv is not None else None,
                (str(g(r, "after coupon price") or g(r, "MAP (Minimum Advertised Price)") or ""))[:32],
                (g(r, "Image link") or "")[:600],
                _f(g(r, "Long")), _f(g(r, "Wide")), _f(g(r, "High")),
                _f(g(r, "Goods Weight")),
            ))

        conn = DBManager.get_connection()
        try:
            with conn.cursor() as cur:
                for i in range(0, len(rows), 500):
                    cur.executemany("""
                        INSERT INTO autooperate.vevor_feed
                            (sku, product_type, title, inventory, price, image,
                             long_in, wide_in, high_in, weight_lb, synced_at)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
                        ON DUPLICATE KEY UPDATE
                            product_type=VALUES(product_type), title=VALUES(title),
                            inventory=VALUES(inventory), price=VALUES(price),
                            image=VALUES(image), long_in=VALUES(long_in),
                            wide_in=VALUES(wide_in), high_in=VALUES(high_in),
                            weight_lb=VALUES(weight_lb), synced_at=NOW()""",
                        rows[i:i + 500])
            conn.commit()
        finally:
            conn.close()
        print(f"vevor_feed synced: {len(rows)} rows")
    return 0


if __name__ == "__main__":
    lp = sys.argv[1] if len(sys.argv) > 1 else None
    sys.exit(main(lp))
