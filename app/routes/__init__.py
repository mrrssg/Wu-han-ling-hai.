"""
✅ v3.3 最兼容版（适配你当前 Bright Data /request 的校验规则）
- 不使用 payload: timeout / wait_until / x-unblock-expect（你这三个都会 400）
- 只使用：zone / url / format(raw) / data_format(screenshot) / render / country / headers
- 截图白图自动重试
依赖：
  pip install requests openpyxl pillow
"""

import os
import re
import time
import random
from pathlib import Path

import requests
from PIL import Image, ImageStat
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


# =================== 你要改的配置 ===================
EXCEL_PATH = r"C:\Users\89415\Desktop\新建 XLSX 工作表 (4).xlsx"
OUT_EXCEL  = r"C:\Users\89415\Desktop\tracking_with_images.xlsx"
SHEET_IDX  = 0

TRACK_COL  = "A"
IMG_COL    = "B"
START_ROW  = 2

SAVE_DIR   = r"C:\Users\89415\Desktop\ups_unlocker_imgs"
# ====================================================

BD_ENDPOINT = "https://api.brightdata.com/request"
BD_ZONE = os.getenv("BRIGHTDATA_ZONE", "unblocked").strip()
BD_API_KEY = os.getenv("BRIGHTDATA_API_KEY", "").strip()

UPS_URL_TMPL = "https://www.ups.com/track?loc=en_US&tracknum={}"

COUNTRY = "us"
RETRY_PER_TN = 5
REQUEST_TIMEOUT = 220       # ✅ 只能在 requests 这里控超时
SLEEP_BETWEEN = (2.0, 5.0)

ROW_HEIGHT = 140
IMG_WIDTH  = 520
IMG_HEIGHT = 130

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
      "AppleWebKit/537.36 (KHTML, like Gecko) "
      "Chrome/126.0.0.0 Safari/537.36")


def ensure_dir(p: str):
    Path(p).mkdir(parents=True, exist_ok=True)

def safe_name(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "_", str(s).strip())[:80]

def normalize_tracking(s: str) -> str:
    return re.sub(r"\s+", "", str(s)).upper()

def insert_image(ws, cell_addr: str, img_path: str):
    img = XLImage(img_path)
    img.width = IMG_WIDTH
    img.height = IMG_HEIGHT
    ws.add_image(img, cell_addr)

def looks_like_png(data: bytes) -> bool:
    return len(data) > 1000 and data[:8] == b"\x89PNG\r\n\x1a\n"

def is_mostly_white(png_path: str, white_mean_threshold: float = 245.0, low_std_threshold: float = 8.0) -> bool:
    im = Image.open(png_path).convert("L")
    stat = ImageStat.Stat(im)
    mean = stat.mean[0]
    std = stat.stddev[0]
    return (mean >= white_mean_threshold) and (std <= low_std_threshold)

def unlocker_screenshot_png(url: str) -> bytes:
    """
    兼容你当前接口：不带 timeout/wait_until/expect
    """
    if not BD_API_KEY:
        raise RuntimeError("未检测到环境变量 BRIGHTDATA_API_KEY（建议先Rotate Key）")

    payload = {
        "zone": BD_ZONE,
        "url": url,
        "format": "raw",                 # ✅ 只能 raw/json
        "data_format": "screenshot",     # ✅ 返回PNG
        "render": True,                  # ✅ 尽量让它渲染
        "country": COUNTRY,
        "headers": {
            "User-Agent": UA,
            "Accept-Language": "en-US,en;q=0.9",
        },
    }

    r = requests.post(
        BD_ENDPOINT,
        headers={
            "Authorization": f"Bearer {BD_API_KEY}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=REQUEST_TIMEOUT,
    )

    if r.status_code != 200:
        raise RuntimeError(f"Unlocker HTTP {r.status_code}: {r.text[:400]}")

    return r.content

def main():
    ensure_dir(SAVE_DIR)

    wb = load_workbook(EXCEL_PATH)
    ws = wb.worksheets[SHEET_IDX]
    max_row = ws.max_row

    ok_cnt = 0
    fail_cnt = 0

    for r in range(START_ROW, max_row + 1):
        raw = ws[f"{TRACK_COL}{r}"].value
        if raw is None:
            continue

        tn = normalize_tracking(raw)
        if not tn:
            continue

        ws.row_dimensions[r].height = ROW_HEIGHT

        url = UPS_URL_TMPL.format(tn)
        out_png = str(Path(SAVE_DIR) / f"{r:05d}_{safe_name(tn)}.png")

        success = False
        last_err = ""

        for attempt in range(RETRY_PER_TN):
            try:
                data = unlocker_screenshot_png(url)

                if not looks_like_png(data):
                    snippet = data[:200].decode("utf-8", errors="ignore")
                    raise RuntimeError(f"返回内容不是PNG：{snippet}")

                with open(out_png, "wb") as f:
                    f.write(data)

                # 判定空白/过小则重试
                size_ok = Path(out_png).stat().st_size >= 15_000
                white = is_mostly_white(out_png)

                if (not size_ok) or white:
                    raise RuntimeError(f"截图疑似空白（size_ok={size_ok}, white={white}）")

                ws[f"{IMG_COL}{r}"].value = None
                insert_image(ws, f"{IMG_COL}{r}", out_png)

                ok_cnt += 1
                print(f"[ROW {r}] {tn} -> OK")
                success = True
                break

            except Exception as e:
                last_err = str(e)
                time.sleep(random.uniform(*SLEEP_BETWEEN) + attempt * 1.2)

        if not success:
            ws[f"{IMG_COL}{r}"].value = f"FAIL: {last_err[:120]}"
            fail_cnt += 1
            print(f"[ROW {r}] {tn} -> FAIL: {last_err}")

        if (ok_cnt + fail_cnt) % 20 == 0:
            time.sleep(random.uniform(8.0, 15.0))

    wb.save(OUT_EXCEL)
    print(f"完成：成功 {ok_cnt}，失败 {fail_cnt}")
    print("输出：", OUT_EXCEL)


if __name__ == "__main__":
    main()
