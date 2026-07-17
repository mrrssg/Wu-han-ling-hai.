# -*- coding: utf-8 -*-
"""产品安全防控 Phase 1（2026-07-17用户拍板）。

链路：录入案例（涉事供应商SKU + PDF/Excel/图片附件）→
  ① 直接命中：供应商SKU → offerprice_listing(4个Mirakl店) + autooperate.mapping_table(其它渠道映射)
     → 各店铺shop SKU，在卖的写 issue_log(safety_risk) 上首页红条
  ② 同款家族：同SKU前缀的变体（GT3021GR+/BL 同模具同风险）自动并入
Phase 2(待做)：AI风险指纹提炼 + 全库举一反三 + 新offer增量夜筛。
"""
import json
import os
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Set

from app.models.db_manager import DBManager

CASE_TYPES = ["侵权", "Prop65", "危险品", "召回", "其它"]
ALLOWED_EXT = {".pdf", ".xlsx", ".xls", ".png", ".jpg", ".jpeg", ".webp"}

_SKU_TOKEN = re.compile(r"^[A-Za-z0-9+\-]{4,30}$")


def _q(conn, sql, params=None):
    with conn.cursor() as cur:
        cur.execute(sql, params) if params else cur.execute(sql)
        return cur.fetchall() or []


def files_root() -> str:
    from flask import current_app
    base = current_app.config.get("BASE_DIR", current_app.root_path)
    d = os.path.join(base, "instance", "safety_cases")
    os.makedirs(d, exist_ok=True)
    return d


def parse_xlsx_skus(path: str) -> Set[str]:
    """从Excel附件里薅出候选SKU token（随后必须过 _validate_supplier_skus）。"""
    from openpyxl import load_workbook
    out: Set[str] = set()
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is None:
                    continue
                s = str(v).strip()
                if _SKU_TOKEN.match(s) and any(ch.isdigit() for ch in s):
                    out.add(s)
    wb.close()
    return out


def _validate_supplier_skus(conn, candidates: Set[str]) -> Set[str]:
    """只留真实存在的供应商SKU（供应商表 ∪ offer表warehouse ∪ 映射表warehouse）。"""
    valid: Set[str] = set()
    cands = [c for c in candidates if c]
    for i in range(0, len(cands), 500):
        chunk = cands[i:i + 500]
        ph = ",".join(["%s"] * len(chunk))
        for sql in (
            f"SELECT SKU AS s FROM autooperate.newestdropship WHERE SKU IN ({ph})",
            f"SELECT SKU AS s FROM autooperate.newestdropship_vevor WHERE SKU IN ({ph})",
            f"SELECT DISTINCT warehouse_sku AS s FROM order_system.offerprice_listing "
            f"WHERE warehouse_sku IN ({ph})",
            f"SELECT DISTINCT warehouse_SKU AS s FROM autooperate.mapping_table "
            f"WHERE warehouse_SKU IN ({ph})",
        ):
            for r in _q(conn, sql, chunk):
                if r["s"]:
                    valid.add(r["s"])
    return valid


def _family_base(sku: str) -> str:
    """GT3021GR+ → GT3021（去尾部颜色后缀）。太短的不冒险截。"""
    s = (sku or "").strip().rstrip("+")
    m = re.match(r"^(.*?\d)[A-Za-z]{0,4}$", s)
    base = m.group(1) if m else s
    return base if len(base) >= 5 else s


def expand_family(conn, skus: Set[str]) -> Dict[str, str]:
    """{家族SKU: 触发它的案例SKU}。同前缀且base一致才算同款（防GT3021误吞GT30215）。"""
    fam: Dict[str, str] = {}
    for sku in skus:
        fam.setdefault(sku, sku)
        base = _family_base(sku)
        like = base + "%"
        for sql in (
            "SELECT SKU AS s FROM autooperate.newestdropship WHERE SKU LIKE %s",
            "SELECT SKU AS s FROM autooperate.newestdropship_vevor WHERE SKU LIKE %s",
            "SELECT DISTINCT warehouse_sku AS s FROM order_system.offerprice_listing "
            "WHERE warehouse_sku LIKE %s",
        ):
            for r in _q(conn, sql, (like,)):
                s2 = r["s"]
                if s2 and _family_base(s2) == base:
                    fam.setdefault(s2, sku)
    return fam


def _resolve_hits(conn, fam: Dict[str, str], direct: Set[str]) -> List[Dict]:
    """家族SKU → 各店铺shop SKU（含停卖的），映射表兜非Mirakl渠道。"""
    hits: List[Dict] = []
    seen_shop: Set[str] = set()
    fam_skus = list(fam)
    for i in range(0, len(fam_skus), 500):
        chunk = fam_skus[i:i + 500]
        ph = ",".join(["%s"] * len(chunk))
        for r in _q(conn, f"""
                SELECT platform, shop_name, shop_sku, warehouse_sku, active
                FROM order_system.offerprice_listing
                WHERE warehouse_sku IN ({ph})""", chunk):
            hits.append({
                "supplier_sku": r["warehouse_sku"],
                "source_sku": fam[r["warehouse_sku"]],
                "hit_type": "direct" if r["warehouse_sku"] in direct else "family",
                "platform": r["platform"], "shop_name": r["shop_name"],
                "shop_sku": r["shop_sku"], "active": int(r["active"] or 0),
            })
            seen_shop.add(r["shop_sku"])
        for r in _q(conn, f"""
                SELECT SKU AS shop_sku, warehouse_SKU AS wsku, owner
                FROM autooperate.mapping_table WHERE warehouse_SKU IN ({ph})""", chunk):
            if r["shop_sku"] in seen_shop:
                continue
            hits.append({
                "supplier_sku": r["wsku"], "source_sku": fam.get(r["wsku"], r["wsku"]),
                "hit_type": "direct" if r["wsku"] in direct else "family",
                "platform": "(映射表)", "shop_name": r["owner"] or "?",
                "shop_sku": r["shop_sku"], "active": None,
            })
            seen_shop.add(r["shop_sku"])

    # 近90天单量（在卖优先级排序用；按offer_sku跨店聚合即可）
    if hits:
        shop_skus = list({h["shop_sku"] for h in hits})
        counts: Dict[str, int] = {}
        for i in range(0, len(shop_skus), 300):
            chunk = shop_skus[i:i + 300]
            ph = ",".join(["%s"] * len(chunk))
            for tbl in ("lowes_order_data", "macy_order_data", "bestbuy_order_data"):
                for r in _q(conn, f"""
                        SELECT offer_sku AS s, COUNT(DISTINCT order_id) AS n
                        FROM order_system.{tbl}
                        WHERE offer_sku IN ({ph}) AND order_state<>'CANCELED'
                          AND created_date >= DATE_SUB(CURDATE(), INTERVAL 90 DAY)
                        GROUP BY offer_sku""", chunk):
                    counts[r["s"]] = counts.get(r["s"], 0) + int(r["n"] or 0)
        for h in hits:
            h["orders_90d"] = counts.get(h["shop_sku"], 0)
    return hits


def create_case(case_type: str, title: str, supplier: str, case_text: str,
                skus_text: str, saved_files: List[Dict]) -> Dict[str, Any]:
    """建案例 + 解析Excel附件里的SKU + 命中解析 + 在卖的写issue_log红条。"""
    typed = {s.strip() for s in re.split(r"[\s,，;；\n]+", skus_text or "") if s.strip()}
    excel_skus: Set[str] = set()
    for f in saved_files:
        if f["path"].lower().endswith((".xlsx", ".xls")):
            try:
                excel_skus |= parse_xlsx_skus(f["path"])
            except Exception:
                pass

    conn = DBManager.get_connection()
    try:
        valid = _validate_supplier_skus(conn, typed | excel_skus)
        invalid_typed = sorted(typed - valid)
        direct = valid
        fam = expand_family(conn, direct) if direct else {}
        hits = _resolve_hits(conn, fam, direct) if fam else []

        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO order_system.safety_case
                    (case_type, title, supplier, case_text, supplier_skus, files_json)
                VALUES (%s,%s,%s,%s,%s,%s)""",
                (case_type, title, supplier, case_text,
                 ",".join(sorted(direct)),
                 json.dumps([{"name": f["name"],
                              "path": os.path.basename(os.path.dirname(f["path"])) + "/" + f["name"]}
                             for f in saved_files], ensure_ascii=False)))
            case_id = cur.lastrowid
            for h in hits:
                cur.execute("""
                    INSERT IGNORE INTO order_system.safety_hit
                        (case_id, supplier_sku, source_sku, hit_type, platform,
                         shop_name, shop_sku, active, orders_90d, reason)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (case_id, h["supplier_sku"], h["source_sku"], h["hit_type"],
                     h["platform"], h["shop_name"], h["shop_sku"], h["active"],
                     h.get("orders_90d", 0),
                     f"案例[{title}] {'涉事SKU' if h['hit_type']=='direct' else '同款家族'}："
                     f"{h['supplier_sku']}"))
            # 在卖的 → 首页红条（issue_log 去重插入）
            for h in hits:
                if h["active"] != 1:
                    continue
                entity = f"{h['shop_sku']}@{h['platform']}-{h['shop_name']}"
                cur.execute("""
                    SELECT id FROM order_system.issue_log
                    WHERE issue_type='safety_risk' AND entity=%s AND status='open'""",
                    (entity,))
                if not cur.fetchone():
                    cur.execute("""
                        INSERT INTO order_system.issue_log
                            (detected_date, issue_type, entity, severity, impact_usd,
                             evidence, suggestion, status)
                        VALUES (CURDATE(), 'safety_risk', %s, 'high', 0, %s,
                                '立即评估下架；处理后在 安全防控 页把该命中标为已下架', 'open')""",
                        (entity,
                         f"安全案例#{case_id}[{case_type}]{title}：供应商SKU "
                         f"{h['supplier_sku']}（{'涉事' if h['hit_type']=='direct' else '同款家族'}）在卖"))
        conn.commit()
    finally:
        conn.close()

    selling = sum(1 for h in hits if h["active"] == 1)
    return {"case_id": case_id, "skus_valid": len(direct),
            "skus_invalid": invalid_typed, "family_size": len(fam),
            "hits": len(hits), "selling": selling}


def mark_hit(hit_id: int, status: str) -> bool:
    if status not in ("delisted", "false_positive", "open"):
        return False
    conn = DBManager.get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM order_system.safety_hit WHERE id=%s", (hit_id,))
            h = cur.fetchone()
            if not h:
                return False
            cur.execute("UPDATE order_system.safety_hit SET status=%s WHERE id=%s",
                        (status, hit_id))
            if status in ("delisted", "false_positive"):
                entity = f"{h['shop_sku']}@{h['platform']}-{h['shop_name']}"
                cur.execute("""
                    UPDATE order_system.issue_log SET status='resolved'
                    WHERE issue_type='safety_risk' AND entity=%s AND status='open'""",
                    (entity,))
        conn.commit()
        return True
    finally:
        conn.close()


# =====================================================================
# Phase 2（2026-07-17）：产品文本缓存 + AI风险指纹 + 全库举一反三扫描
# =====================================================================

SCAN_MAX_CANDIDATES = 150     # 粗筛后最多送AI精判的数量
FP_MAX_EXAMPLE_SKUS = 3       # 提炼指纹时最多看几个涉事SKU的资料


def _catalog_skus(conn) -> Set[str]:
    """我们目录里的全部供应商SKU（offer表 ∪ 映射表）。"""
    out: Set[str] = set()
    for sql in (
        "SELECT DISTINCT warehouse_sku AS s FROM order_system.offerprice_listing "
        "WHERE warehouse_sku IS NOT NULL AND warehouse_sku<>''",
        "SELECT DISTINCT warehouse_SKU AS s FROM autooperate.mapping_table",
    ):
        for r in _q(conn, sql):
            if r["s"]:
                out.add(r["s"])
    return out


# Costway 官方全量 dropship feed（用户指示直接下载同步，不绕飞书）。
# WinZip AES-256 加密，密码会轮换——都失败时要向用户要新密码加到列表最前。
FEED_URL = "https://cdn.costway.com/media/feed/dropship.zip"
FEED_PASSWORDS = ["728465", "482619", "915473"]


def _sync_costway_from_feed(catalog: Set[str]) -> List[tuple]:
    """下载feed→pyzipper解压→CSV→行元组。33k+ SKU全量缓存（含我们没卖的，
    in_catalog=0，留给选品前预警用）。"""
    import csv as _csv
    import io as _io
    import requests
    import pyzipper

    r = requests.get(FEED_URL, timeout=300)
    r.raise_for_status()
    buf = _io.BytesIO(r.content)
    data = None
    last_exc: Optional[Exception] = None
    for pwd in FEED_PASSWORDS:
        try:
            buf.seek(0)
            with pyzipper.AESZipFile(buf) as zf:
                zf.setpassword(pwd.encode())
                name = next(n for n in zf.namelist() if n.lower().endswith(".csv"))
                data = zf.read(name)
            break
        except Exception as exc:
            last_exc = exc
    if data is None:
        raise RuntimeError(f"feed解压失败——密码可能又轮换了（试过{FEED_PASSWORDS}），"
                           f"要向老板要新密码: {last_exc}")

    rows = []
    reader = _csv.DictReader(_io.StringIO(data.decode("utf-8", errors="replace")))
    for rec in reader:
        sku = (rec.get("SKU") or "").strip()
        if not sku:
            continue
        rows.append(("Costway", sku,
                     (rec.get("Item Name") or "")[:400],
                     (rec.get("Specification") or "")[:1200],
                     (rec.get("Description") or "")[:1200],
                     (rec.get("Images1") or "")[:600],
                     (rec.get("Category") or "")[:160],
                     1 if sku in catalog else 0))
    return rows


def _sync_vevor_from_feishu(catalog: Set[str]) -> List[tuple]:
    """司顺没有feed，仍从飞书sku资料表翻页拉（只留目录内的）。"""
    import requests
    from app.services.listing_sentinel_service import (
        VEVOR_TBL, PRODUCT_APP, _token, _gt, _glink)

    headers = {"Authorization": f"Bearer {_token()}",
               "Content-Type": "application/json"}
    rows = []
    page_token = ""
    while True:
        url = (f"https://open.feishu.cn/open-apis/bitable/v1/apps/{PRODUCT_APP}"
               f"/tables/{VEVOR_TBL}/records?page_size=500"
               + (f"&page_token={page_token}" if page_token else ""))
        r = requests.get(url, headers=headers, timeout=60).json()
        data = r.get("data") or {}
        for item in data.get("items") or []:
            f = item.get("fields") or {}
            sku = _gt(f.get("SKU")).strip()
            if not sku or sku not in catalog:
                continue
            sells = " / ".join(_gt(f.get(f"Selling point {i}")) for i in range(1, 6)
                               if _gt(f.get(f"Selling point {i}")))
            rows.append(("Vevor", sku,
                         _gt(f.get("Product title"))[:400], sells[:1200],
                         _gt(f.get("Product description"))[:1200],
                         _glink(f.get("Image link"))[:600], None, 1))
        if not data.get("has_more"):
            break
        page_token = data.get("page_token") or ""
        if not page_token:
            break
    return rows


def sync_product_cache() -> Dict[str, Any]:
    """豪雅=官方feed直下（全量33k+，含未上架的留给选品预警）；司顺=飞书。"""
    conn = DBManager.get_connection()
    try:
        catalog = _catalog_skus(conn)
    finally:
        conn.close()

    summary: Dict[str, Any] = {"catalog": len(catalog)}
    try:
        c_rows = _sync_costway_from_feed(catalog)
        summary["costway_source"] = "feed"
    except Exception as exc:
        summary["costway_source"] = f"feed失败:{str(exc)[:160]}"
        c_rows = []
    try:
        v_rows = _sync_vevor_from_feishu(catalog)
    except Exception as exc:
        summary["vevor_error"] = str(exc)[:160]
        v_rows = []
    rows = c_rows + v_rows

    conn = DBManager.get_connection()
    try:
        with conn.cursor() as cur:
            for i in range(0, len(rows), 300):
                cur.executemany("""
                    INSERT INTO order_system.safety_product_cache
                        (supplier, sku, title, spec, descr, image_url, category,
                         in_catalog, synced_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,NOW())
                    ON DUPLICATE KEY UPDATE title=VALUES(title), spec=VALUES(spec),
                        descr=VALUES(descr), image_url=VALUES(image_url),
                        category=VALUES(category), in_catalog=VALUES(in_catalog),
                        synced_at=NOW()""", rows[i:i + 300])
        conn.commit()
    finally:
        conn.close()
    summary["costway"] = len(c_rows)
    summary["vevor"] = len(v_rows)
    summary["in_catalog"] = sum(1 for r in rows if r[7] == 1)
    return summary


def cache_stats() -> Dict[str, Any]:
    conn = DBManager.get_connection()
    try:
        r = _q(conn, """SELECT COUNT(*) AS n,
                        SUM(CASE WHEN in_catalog=1 THEN 1 ELSE 0 END) AS in_cat,
                        MAX(synced_at) AS t
                        FROM order_system.safety_product_cache""")[0]
        return {"n": int(r["n"] or 0), "in_catalog": int(r["in_cat"] or 0),
                "synced_at": r["t"]}
    finally:
        conn.close()


def _case_row(conn, case_id: int) -> Optional[Dict]:
    rows = _q(conn, "SELECT * FROM order_system.safety_case WHERE id=%s", (case_id,))
    return rows[0] if rows else None


def generate_fingerprint(case_id: int) -> Dict[str, Any]:
    """AI读案例+涉事产品资料 → 结构化风险指纹（用户确认后才用于扫描）。"""
    from app.services.listing_sentinel_service import _openai_client, MODEL_NAME

    conn = DBManager.get_connection()
    try:
        case = _case_row(conn, case_id)
        if not case:
            return {"success": False, "msg": "案例不存在"}
        skus = [s for s in (case["supplier_skus"] or "").split(",") if s][:FP_MAX_EXAMPLE_SKUS]
        examples = []
        for sku in skus:
            rows = _q(conn, """SELECT title, spec, descr FROM order_system.safety_product_cache
                               WHERE sku=%s LIMIT 1""", (sku,))
            if rows:
                examples.append({"sku": sku, **rows[0]})
    finally:
        conn.close()

    prompt = f"""你是电商合规审计专家。下面是一起产品安全案例，请提炼出用于"举一反三排查相似产品"的结构化风险指纹。

【案例类型】{case['case_type']}
【案例标题】{case['title']}
【案情描述】{case['case_text'] or '（无）'}
【涉事产品资料】{json.dumps(examples, ensure_ascii=False)[:6000] if examples else '（缓存里没有涉事SKU的资料，只按案情描述提炼）'}

按案例类型抓重点：侵权→外观形态/品牌词/独特设计元素；Prop65→材质化学成分(如PVC/黄铜/铅/邻苯)；危险品→危险部件成分(如锂电池/压力罐/易燃物)；召回→缺陷结构。
只输出JSON：
{{"case_essence":"一句话案由本质",
"key_features":["产品关键特征,3-8条"],
"materials":["涉险材质/成分,没有则空数组"],
"brand_terms":["涉险品牌词/型号词,没有则空数组"],
"category_scope":["可能涉及的产品类目,英文"],
"keywords_en":["用于全库粗筛的英文关键词,8-20个,含同义词"],
"keywords_cn":["中文关键词,可选"],
"judge_focus":"精判时应重点核对什么(一句话)"}}"""

    client = _openai_client()
    resp = client.chat.completions.create(
        model=MODEL_NAME, response_format={"type": "json_object"},
        messages=[{"role": "system", "content": "Output valid JSON only."},
                  {"role": "user", "content": prompt}])
    fp = json.loads(resp.choices[0].message.content)

    conn = DBManager.get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""UPDATE order_system.safety_case
                           SET fingerprint_json=%s, scan_status='fingerprint_ready'
                           WHERE id=%s""",
                        (json.dumps(fp, ensure_ascii=False), case_id))
        conn.commit()
    finally:
        conn.close()
    return {"success": True, "fingerprint": fp}


def save_fingerprint(case_id: int, fp_text: str) -> Dict[str, Any]:
    try:
        fp = json.loads(fp_text)
    except ValueError as exc:
        return {"success": False, "msg": f"JSON格式错误: {exc}"}
    conn = DBManager.get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""UPDATE order_system.safety_case
                           SET fingerprint_json=%s, scan_status='fingerprint_ready'
                           WHERE id=%s""",
                        (json.dumps(fp, ensure_ascii=False), case_id))
        conn.commit()
    finally:
        conn.close()
    return {"success": True}


def _stage1_candidates(conn, case: Dict, fp: Dict) -> List[Dict]:
    """粗筛：关键词在缓存文本里打分，排除已命中的家族SKU，取前N。零AI成本。"""
    kws = [k.strip().lower() for k in
           (fp.get("keywords_en") or []) + (fp.get("keywords_cn") or [])
           + (fp.get("brand_terms") or []) + (fp.get("materials") or []) if k and len(k) >= 3]
    if not kws:
        return []
    already = {r["supplier_sku"] for r in _q(conn,
        "SELECT DISTINCT supplier_sku FROM order_system.safety_hit WHERE case_id=%s",
        (case["id"],))}
    scored = []
    for r in _q(conn, "SELECT supplier, sku, title, spec, descr, image_url "
                      "FROM order_system.safety_product_cache WHERE in_catalog=1"):
        if r["sku"] in already:
            continue
        text = " ".join(filter(None, [r["title"], r["spec"], r["descr"]])).lower()
        if not text:
            continue
        score = sum(1 for k in kws if k in text)
        if score > 0:
            r["_score"] = score
            scored.append(r)
    scored.sort(key=lambda x: -x["_score"])
    return scored[:SCAN_MAX_CANDIDATES]


def _judge_candidate(client, model: str, case: Dict, fp: Dict, cand: Dict) -> Dict:
    prompt = f"""你是电商合规审计专家。已知一起{case['case_type']}安全案例，判断下面这个产品是否可能有同类风险。

【案由本质】{fp.get('case_essence','')}
【关键特征】{fp.get('key_features')}
【涉险材质/成分】{fp.get('materials')}
【涉险品牌词】{fp.get('brand_terms')}
【精判重点】{fp.get('judge_focus','')}

【待判产品 {cand['sku']}】
标题: {(cand['title'] or '')[:300]}
规格: {(cand['spec'] or '')[:1200]}
描述: {(cand['descr'] or '')[:1200]}

原则：只有产品资料里有明确依据才判 high；特征部分吻合判 mid；仅类目沾边判 low；无关判 none。宁可漏判不可乱判。
只输出JSON：{{"risk":"high|mid|low|none","reason":"一句话中文理由","evidence":"引用产品资料原文(≤200字)"}}"""
    content = [{"type": "text", "text": prompt}]
    if case["case_type"] == "侵权" and cand.get("image_url"):
        content.append({"type": "image_url", "image_url": {"url": cand["image_url"]}})
    resp = client.chat.completions.create(
        model=model, response_format={"type": "json_object"},
        messages=[{"role": "system", "content": "Output valid JSON only."},
                  {"role": "user", "content": content}])
    return json.loads(resp.choices[0].message.content)


def run_scan(case_id: int) -> Dict[str, Any]:
    """全库举一反三（同步执行，路由用后台线程包）。"""
    from app.services.listing_sentinel_service import _openai_client, MODEL_NAME

    conn = DBManager.get_connection()
    try:
        case = _case_row(conn, case_id)
        if not case or not case.get("fingerprint_json"):
            return {"success": False, "msg": "先生成并确认风险指纹"}
        fp = json.loads(case["fingerprint_json"])
        with conn.cursor() as cur:
            cur.execute("UPDATE order_system.safety_case SET scan_status='scanning' "
                        "WHERE id=%s", (case_id,))
        conn.commit()
        cands = _stage1_candidates(conn, case, fp)
    finally:
        conn.close()

    client = _openai_client()
    stats = {"candidates": len(cands), "judged": 0, "high": 0, "mid": 0, "low": 0}
    for cand in cands:
        try:
            verdict = _judge_candidate(client, MODEL_NAME, case, fp, cand)
        except Exception:
            continue
        stats["judged"] += 1
        risk = (verdict.get("risk") or "none").lower()
        if risk not in ("high", "mid", "low"):
            continue
        stats[risk] += 1
        conn = DBManager.get_connection()
        try:
            fam = {cand["sku"]: cand["sku"]}
            hits = _resolve_hits(conn, fam, set())
            with conn.cursor() as cur:
                for h in hits:
                    cur.execute("""
                        INSERT IGNORE INTO order_system.safety_hit
                            (case_id, supplier_sku, source_sku, hit_type, platform,
                             shop_name, shop_sku, active, orders_90d, risk_level,
                             reason, evidence)
                        VALUES (%s,%s,%s,'ai',%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (case_id, h["supplier_sku"], cand["sku"],
                         h["platform"], h["shop_name"], h["shop_sku"], h["active"],
                         h.get("orders_90d", 0), risk,
                         (verdict.get("reason") or "")[:500],
                         (verdict.get("evidence") or "")[:990]))
                    if risk == "high" and h["active"] == 1:
                        entity = f"{h['shop_sku']}@{h['platform']}-{h['shop_name']}"
                        cur.execute("""SELECT id FROM order_system.issue_log
                                       WHERE issue_type='safety_risk' AND entity=%s
                                         AND status='open'""", (entity,))
                        if not cur.fetchone():
                            cur.execute("""
                                INSERT INTO order_system.issue_log
                                    (detected_date, issue_type, entity, severity, impact_usd,
                                     evidence, suggestion, status)
                                VALUES (CURDATE(), 'safety_risk', %s, 'high', 0, %s,
                                        '立即评估下架；处理后在安全防控页标记', 'open')""",
                                (entity,
                                 f"安全案例#{case_id} AI相似判定(high)：{cand['sku']}——"
                                 f"{(verdict.get('reason') or '')[:300]}"))
            conn.commit()
        finally:
            conn.close()

    conn = DBManager.get_connection()
    try:
        with conn.cursor() as cur:
            stats["done_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cur.execute("""UPDATE order_system.safety_case
                           SET scan_status='done', scan_summary_json=%s WHERE id=%s""",
                        (json.dumps(stats, ensure_ascii=False), case_id))
        conn.commit()
    finally:
        conn.close()
    return {"success": True, **stats}


def close_case(case_id: int) -> None:
    conn = DBManager.get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE order_system.safety_case SET status='closed' WHERE id=%s",
                        (case_id,))
            cur.execute("""
                UPDATE order_system.issue_log i
                JOIN order_system.safety_hit h
                  ON i.entity = CONCAT(h.shop_sku,'@',h.platform,'-',h.shop_name)
                SET i.status='resolved'
                WHERE h.case_id=%s AND i.issue_type='safety_risk' AND i.status='open'""",
                (case_id,))
        conn.commit()
    finally:
        conn.close()
