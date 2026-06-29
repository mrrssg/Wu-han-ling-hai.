# =========================
# 3) services/export_service.py
# =========================
# 业务层：查数据 + 调 exporter + 产出(文件流, 文件名)
# 这里不写 Flask send_file，不写路由

from datetime import datetime
import os
from flask import current_app
from app.models.db_manager import DBManager
from app.services import blacklist_service
from io import BytesIO
from typing import List, Dict, Tuple
import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy as xl_copy

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExportService:
    # supplier key -> 供应商价格表（也用于筛 SKU 归属哪个供应商）
    SUPPLIER_TABLE = {
        "haoya": "newestdropship",
        "sishun": "newestdropship_vevor",
        "dajian": "newestdropship_dajian",
    }
    SUPPLIER_LABEL = {"haoya": "豪雅", "sishun": "司顺", "dajian": "大建"}

    @staticmethod
    def _get_supplier_orders(supplier: str) -> Tuple[List[Dict], Dict]:
        """取该供应商的全部未发货订单（按 SKU 归属过滤）+ 价格表。"""
        table = ExportService.SUPPLIER_TABLE[supplier]
        all_orders: List[Dict] = DBManager.fetch_unshipped_orders()
        price_map: Dict = DBManager.fetch_sku_map(table)
        orders = [o for o in all_orders
                  if (o.get('sku') or '').strip() and (o.get('sku') or '').strip() in price_map]
        return orders, price_map

    @staticmethod
    def screen_supplier(supplier: str) -> Dict:
        """筛查：返回全量 / 正常 / 命中黑名单，用于筛查结果页。"""
        orders, price_map = ExportService._get_supplier_orders(supplier)
        clean, hits = blacklist_service.screen_orders(orders)
        return {"supplier": supplier, "label": ExportService.SUPPLIER_LABEL[supplier],
                "total": len(orders), "clean": clean, "hits": hits, "price_map": price_map}

    @staticmethod
    def export_intercept_xlsx(supplier: str) -> Tuple[BytesIO, str]:
        """命中黑名单的订单清单（供人工复核，不发供应商）。"""
        res = ExportService.screen_supplier(supplier)
        bio = ExportService.build_intercept_xlsx(res["hits"], res["label"])
        filename = (f"{res['label']}_黑名单拦截_"
                    f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        return bio, filename

    @staticmethod
    def build_intercept_xlsx(hits: List[Dict], label: str) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "黑名单拦截"
        headers = ["供应商", "命中维度", "拉黑原因", "平台单号", "Costway单号", "SKU",
                   "数量", "收货人", "电话", "邮箱", "街道", "城市", "州", "邮编"]
        ws.append(headers)
        for o in hits:
            full_name = (o.get("first_name") or "").strip()
            if o.get("last_name"):
                full_name = f"{full_name} {o.get('last_name')}".strip()
            ws.append([
                label,
                "、".join(o.get("_bl_matched_on") or []),
                o.get("_bl_reason") or "",
                o.get("order_number") or "",
                o.get("costway_number") or "",
                (o.get("sku") or "").strip(),
                o.get("qty") or 1,
                full_name,
                o.get("phone") or "",
                o.get("customer_email") or "",
                o.get("street") or "",
                o.get("city") or "",
                o.get("region") or "",
                o.get("postcode") or "",
            ])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        wb.close()
        return bio

    @staticmethod
    def export_haoya_unshipped_xlsx() -> Tuple[BytesIO, str]:
        # 取豪雅订单 -> 筛掉黑名单 -> 只导出正常订单
        haoya_orders, haoya_price_map = ExportService._get_supplier_orders('haoya')
        clean, _hits = blacklist_service.screen_orders(haoya_orders)

        bio = ExportService.build_haoya_xlsx(clean, haoya_price_map)
        filename = f"豪雅_未发货_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return bio, filename

    @staticmethod
    def build_haoya_xlsx(orders: List[Dict], price_map: Dict) -> BytesIO:
        # ✅ 模板路径（你按实际路径改）
        template_path = os.path.join(
            current_app.root_path, "static", "templates", "haoya_template.xlsx"
        )

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active  # 模板一般只有一个sheet；如果不是，改成 wb["xxx"]

        # ✅ 从第2行开始写（第1行是模板表头）
        start_row = 2

        def calc_amount(sku: str) -> float:
            origin = price_map.get(sku, 0)
            try:
                origin = float(origin) if origin else 0
            except Exception:
                origin = 0
            return origin * 0.75 if origin else 0

        for i, o in enumerate(orders):
            row = start_row + i

            sku = (o.get("sku") or "").strip()
            qty = o.get("qty") or 1
            amount = calc_amount(sku)

            full_name = (o.get("first_name") or "").strip()
            if o.get("last_name"):
                full_name = f"{full_name} {o.get('last_name')}".strip()

            # ⚠️ 下面这些“写入到哪一列”，必须和你的模板列一致
            # 我按你截图里常见字段给一个映射，你只需要调整列字母即可。

            # A 订单编号（截图里像 WHLH260106-8 这种）
            ws[f"A{row}"] = o.get("costway_number") or ""

            # B 邮箱
            ws[f"B{row}"] = o.get("customer_email") or ""

            # C 交易时间（留空）
            ws[f"C{row}"] = ""

            # D 收货人国家中文（留空）
            ws[f"D{row}"] = ""

            # E 币种（留空）
            ws[f"E{row}"] = ""

            # F 订单金额（留空）
            ws[f"F{row}"] = ""

            # G 收货国家(英文)（留空）
            ws[f"G{row}"] = ""

            # H 收货国家代码（留空）
            ws[f"H{row}"] = ""

            # I 收货人
            ws[f"I{row}"] = full_name

            # J 收货人地址
            ws[f"J{row}"] = o.get("street") or ""

            # K 收货人地址2（你DB如果没拆出来，就留空或从 o.get("street2")）
            ws[f"K{row}"] = o.get("street2") or o.get("address2") or ""

            # L 收货城市
            ws[f"L{row}"] = o.get("city") or ""

            # M 州或者省
            ws[f"M{row}"] = o.get("region") or ""

            # N 收货邮编
            ws[f"N{row}"] = o.get("postcode") or ""

            # O 收货人电话
            ws[f"O{row}"] = o.get("phone") or "0000000000"

            # P 付款人（模板可能固定空）
            ws[f"P{row}"] = ""

            # Q SKU
            ws[f"Q{row}"] = sku

            # R 中文名（你没有就空）
            ws[f"R{row}"] = ""

            # S 商品盒（留空）
            ws[f"S{row}"] = ""

            # T 英文名（没有就空）
            ws[f"T{row}"] = ""

            # U 数量
            ws[f"U{row}"] = qty

            # V 重量KG（你先写1）
            ws[f"V{row}"] = ""

            # W 报关价（先空）
            ws[f"W{row}"] = ""

            # X 原产国（先空）
            ws[f"X{row}"] = ""

            # Y 备注
            ws[f"Y{row}"] = ""

            # Z 商品名
            ws[f"Z{row}"] = ""

            # AA 销售单号
            ws[f"AA{row}"] = o.get("sales_no") or ""

            # AB 店铺
            ws[f"AB{row}"] = o.get("store") or "FDSUS-BP"

            # AC 付货方
            ws[f"AC{row}"] = ""

            # AD 承运
            ws[f"AD{row}"] = ""

            # AE Trans
            ws[f"AE{row}"] = ""

            # AF ionID（平台单号/Costway单号）
            ws[f"AF{row}"] = o.get("order_number") or ""

        # ✅ 保存到内存流（用于 send_file 下载）
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        wb.close()
        return bio

    @staticmethod
    def export_sishun_unshipped_xlsx() -> Tuple[BytesIO, str]:
        # 取司顺订单 -> 筛掉黑名单 -> 只导出正常订单
        sishun_orders, sishun_price_map = ExportService._get_supplier_orders('sishun')
        clean, _hits = blacklist_service.screen_orders(sishun_orders)

        bio = ExportService.build_sishun_xlsx(clean, sishun_price_map)
        filename = f"司顺_未发货_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return bio, filename

    @staticmethod
    def export_dajian_unshipped_xls() -> Tuple[BytesIO, str]:
        # 取大建订单 -> 筛掉黑名单 -> 只导出正常订单
        dajian_orders, _dajian_price_map = ExportService._get_supplier_orders('dajian')
        clean, _hits = blacklist_service.screen_orders(dajian_orders)

        bio = ExportService.build_dajian_xls(clean)
        filename = f"大建_未发货_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
        return bio, filename

    @staticmethod
    def build_dajian_xls(orders: List[Dict]) -> BytesIO:
        template_path = os.path.join(
            current_app.root_path, "static", "templates", "OrderTemplateUS(29).xls"
        )
        book = xlrd.open_workbook(template_path, formatting_info=True)
        sheet = book.sheet_by_index(0)

        # 读取模板表头（第1行字段名），保持原样式
        headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
        header_map = {h: idx for idx, h in enumerate(headers)}

        wb = xl_copy(book)
        ws = wb.get_sheet(0)

        def write_cell(row_idx: int, key: str, value):
            if key not in header_map:
                return
            col_idx = header_map[key]
            ws.write(row_idx, col_idx, value)

        start_row = 2  # 保留模板前两行（字段名+说明）

        for i, o in enumerate(orders, start=0):
            row_idx = start_row + i
            full_name = (o.get("first_name") or "").strip()
            if o.get("last_name"):
                full_name = f"{full_name} {o.get('last_name')}".strip()

            order_date = o.get("order_date") or ""
            if hasattr(order_date, "strftime"):
                order_date = order_date.strftime("%Y-%m-%d")

            write_cell(row_idx, "ShipFrom", o.get("costway_number") or "")
            write_cell(row_idx, "*OrderId", o.get("order_number") or "")
            write_cell(row_idx, "*LineItemNumber", o.get("line_item_number") or "")
            write_cell(row_idx, "*B2BItemCode", o.get("sku") or "")
            write_cell(row_idx, "*ShipToQty", o.get("qty") or 1)
            write_cell(row_idx, "DeliveryService", "")
            write_cell(row_idx, "*ShipToName", full_name)
            write_cell(row_idx, "ShipToEmail", "kevin@ecooso.com")
            write_cell(row_idx, "*ShipToPhone", o.get("phone") or "0000000000")
            write_cell(row_idx, "*AddressLine1", o.get("street") or "")
            write_cell(row_idx, "AddressLine2", o.get("street2") or o.get("address2") or "")
            write_cell(row_idx, "*ShipToCity", o.get("city") or "")
            write_cell(row_idx, "*ShipToState", o.get("region") or "")
            write_cell(row_idx, "*ShipToPostalCode", o.get("postcode") or "")
            write_cell(row_idx, "*ShipToCountry", "US")
            write_cell(row_idx, "ShipToServiceLevel", "")
            write_cell(row_idx, "ShipToAttachmentUrl", "")
            write_cell(row_idx, "OrderDate", order_date)
            write_cell(row_idx, "BuyerBrand", "")
            write_cell(row_idx, "BuyerPlatformSku", o.get("platform_sku") or "")
            write_cell(row_idx, "BuyerSkuDescription", "")
            write_cell(row_idx, "BuyerSkuCommercialValue", "")
            write_cell(row_idx, "BuyerSkuLink", "")
            write_cell(row_idx, "OrderComments", "")

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    @staticmethod
    def build_sishun_xlsx(orders: List[Dict], price_map: Dict) -> BytesIO:
        # ✅ 模板路径（你按实际路径改）
        template_path = os.path.join(
            current_app.root_path, "static", "templates", "sishun_template.xlsx"
        )

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active  # 模板一般只有一个sheet；如果不是，改成 wb["xxx"]

        # ✅ 从第2行开始写（第1行是模板表头）
        start_row = 2

        def calc_amount(sku: str) -> float:
            origin = price_map.get(sku, 0)
            try:
                origin = float(origin) if origin else 0
            except Exception:
                origin = 0
            return origin * 1 if origin else 0

        for i, o in enumerate(orders):
            row = start_row + i

            sku = (o.get("sku") or "").strip()
            qty = o.get("qty") or 1
            amount = calc_amount(sku)
            store_key = (o.get("store_key") or "").strip().lower()

            full_name = (o.get("first_name") or "").strip()
            if o.get("last_name"):
                full_name = f"{full_name} {o.get('last_name')}".strip()

            # ⚠️ 下面这些“写入到哪一列”，必须和你的模板列一致
            # 我按你截图里常见字段给一个映射，你只需要调整列字母即可。

            # A 订单编号（截图里像 WHLH260106-8 这种）
            ws[f"A{row}"] = o.get("order_number") or ""

            # 第三方订单号（选填）
            ws[f"B{row}"] = qty

            # C 交易时间（没有就用今天）
            ws[f"C{row}"] = amount

            # D 收货人国家中文
            ws[f"D{row}"] = sku

            # E 币种
            ws[f"E{row}"] = o.get("first_name")


            ws[f"F{row}"] = o.get("last_name")

            # G 收货国家(英文)
            ws[f"G{row}"] = o.get("phone") or "0000000000"

            # H 收货国家代码
            ship_email_map = {
                "macy_kuyotq": "kuyotqmc@hotmail.com",
                "macy_wopet": "wopetmc@hotmail.com",
            }
            ws[f"H{row}"] = ship_email_map.get(store_key, "kuyotqmc@hotmail.com")

            # I 收货人
            ws[f"I{row}"] = o.get("postcode") or ""

            # J 收货人地址
            ws[f"J{row}"] = "United States"

            # K 收货人地址2（你DB如果没拆出来，就留空或从 o.get("street2")）
            ws[f"K{row}"] = "US"

            # L 收货城市
            ws[f"L{row}"] = o.get("region") or ""

            # M 州或者省
            ws[f"M{row}"] = o.get("city") or ""

            ws[f"N{row}"] = o.get("street") or ""

            # O 收货人电话
            ws[f"O{row}"] = ""

            ws[f"P{row}"] = "wuhanlinghai@163.com"

            # Q CostwayOrder (WHLH...)
            ws[f"Q{row}"] = o.get("costway_number") or ""



        # ✅ 保存到内存流（用于 send_file 下载）
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        wb.close()
        return bio
