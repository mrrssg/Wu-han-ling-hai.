"""
客户黑名单筛查 (customer blacklist screening) for supplier order exports.

Used by ExportOrder to pull blacklisted orders out of the 豪雅/司顺/大建
unshipped-order exports before the file is sent to the supplier for dropship.

Matching rules (normalised exact match, agreed with user 2026-06-29):
  - phone  (last 10 digits)        -> single hit blocks
  - email  (lowercased)            -> single hit blocks
  - address (street_norm + zip5)   -> single hit blocks
  - name   ALONE never blocks; name must also match zip OR city (avoids
    false-positives on common names like "John Smith")

Order dict fields (from DBManager.fetch_unshipped_orders / ExportOrder):
    first_name, last_name, customer_email, phone, street, city, region(state),
    postcode(zip)

Everything is stored raw in order_system.customer_blacklist and normalised here
at match time; the blacklist is small so we load it fully into memory per run.
"""
import re
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from app.models.db_manager import DBManager


# Excel import / manual-form columns (one source of truth for both paths).
BLACKLIST_COLUMNS = ["姓名", "电话", "邮箱", "街道", "城市", "州", "邮编", "拉黑原因"]

# map Chinese column -> db field
_COL_TO_FIELD = {
    "姓名": "full_name", "电话": "phone", "邮箱": "email", "街道": "street",
    "城市": "city", "州": "state", "邮编": "zip", "拉黑原因": "reason",
}

# common US street-suffix abbreviations -> canonical, so "123 Main St" matches
# "123 Main Street".
_ADDR_ABBR = {
    "street": "st", "avenue": "ave", "road": "rd", "boulevard": "blvd",
    "drive": "dr", "lane": "ln", "court": "ct", "place": "pl", "terrace": "ter",
    "circle": "cir", "parkway": "pkwy", "highway": "hwy", "suite": "ste",
    "apartment": "apt", "north": "n", "south": "s", "east": "e", "west": "w",
}


# =============================================================================
# Normalisation helpers
# =============================================================================

def norm_phone(v) -> str:
    if not v:
        return ""
    digits = re.sub(r"\D", "", str(v))
    return digits[-10:] if len(digits) >= 10 else ""  # need a full US number


def norm_email(v) -> str:
    return str(v).strip().lower() if v else ""


def norm_text(v) -> str:
    return re.sub(r"\s+", " ", str(v).strip().lower()) if v else ""


def norm_zip(v) -> str:
    if not v:
        return ""
    digits = re.sub(r"\D", "", str(v))
    return digits[:5] if len(digits) >= 5 else ""


def norm_addr(v) -> str:
    """Lowercase, drop punctuation, collapse abbreviations + spaces."""
    if not v:
        return ""
    s = str(v).lower()
    s = re.sub(r"[.,#]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    tokens = [_ADDR_ABBR.get(t, t) for t in s.split(" ")]
    return " ".join(t for t in tokens if t)


def _full_name(first, last) -> str:
    name = (first or "").strip()
    if last:
        name = f"{name} {str(last).strip()}".strip()
    return name


# =============================================================================
# Blacklist load + index
# =============================================================================

class _BlacklistIndex:
    """In-memory lookup built from active blacklist rows."""

    def __init__(self, rows: List[Dict]):
        self.phones: Dict[str, Dict] = {}
        self.emails: Dict[str, Dict] = {}
        self.addrs: Dict[Tuple[str, str], Dict] = {}       # (street_norm, zip5)
        self.name_zip: Dict[Tuple[str, str], Dict] = {}    # (name_norm, zip5)
        self.name_city: Dict[Tuple[str, str], Dict] = {}   # (name_norm, city_norm)
        for r in rows:
            p = norm_phone(r.get("phone"))
            if p:
                self.phones.setdefault(p, r)
            e = norm_email(r.get("email"))
            if e:
                self.emails.setdefault(e, r)
            street = norm_addr(r.get("street"))
            z = norm_zip(r.get("zip"))
            city = norm_text(r.get("city"))
            name = norm_text(r.get("full_name"))
            if street and z:
                self.addrs.setdefault((street, z), r)
            if name and z:
                self.name_zip.setdefault((name, z), r)
            if name and city:
                self.name_city.setdefault((name, city), r)

    def match(self, order: Dict) -> Optional[Dict]:
        """Return {'matched_on': [...], 'reason': str, 'entry_id': id} or None."""
        name = norm_text(_full_name(order.get("first_name"), order.get("last_name")))
        phone = norm_phone(order.get("phone"))
        email = norm_email(order.get("customer_email"))
        street = norm_addr(order.get("street"))
        z = norm_zip(order.get("postcode"))
        city = norm_text(order.get("city"))

        matched_on: List[str] = []
        hit_entry: Optional[Dict] = None

        def take(entry, label):
            nonlocal hit_entry
            matched_on.append(label)
            if hit_entry is None:
                hit_entry = entry

        if phone and phone in self.phones:
            take(self.phones[phone], "电话")
        if email and email in self.emails:
            take(self.emails[email], "邮箱")
        if street and z and (street, z) in self.addrs:
            take(self.addrs[(street, z)], "地址")
        # name alone never blocks: require zip OR city alongside the name
        if name and z and (name, z) in self.name_zip:
            take(self.name_zip[(name, z)], "姓名+邮编")
        if name and city and (name, city) in self.name_city:
            take(self.name_city[(name, city)], "姓名+城市")

        if not matched_on:
            return None
        return {
            "matched_on": matched_on,
            "reason": (hit_entry or {}).get("reason") or "",
            "entry_id": (hit_entry or {}).get("id"),
        }


def load_index() -> _BlacklistIndex:
    return _BlacklistIndex(list_entries(active_only=True))


# =============================================================================
# Screening
# =============================================================================

def screen_orders(orders: List[Dict]) -> Tuple[List[Dict], List[Dict]]:
    """Split orders into (clean, hits). Each hit is the order dict plus
    `_bl_matched_on` (list) and `_bl_reason` (str). If the blacklist is empty
    nothing is filtered (clean == orders)."""
    idx = load_index()
    if not (idx.phones or idx.emails or idx.addrs or idx.name_zip or idx.name_city):
        return list(orders), []
    clean, hits = [], []
    for o in orders:
        m = idx.match(o)
        if m is None:
            clean.append(o)
        else:
            h = dict(o)
            h["_bl_matched_on"] = m["matched_on"]
            h["_bl_reason"] = m["reason"]
            hits.append(h)
    return clean, hits


# =============================================================================
# CRUD
# =============================================================================

def list_entries(active_only: bool = False) -> List[Dict]:
    conn = DBManager.get_connection()
    try:
        with conn.cursor() as cursor:
            sql = ("SELECT id, full_name, phone, email, street, city, state, zip, "
                   "reason, active, source, created_by, created_at "
                   "FROM order_system.customer_blacklist")
            if active_only:
                sql += " WHERE active=1"
            sql += " ORDER BY created_at DESC, id DESC"
            cursor.execute(sql)
            return cursor.fetchall() or []
    finally:
        conn.close()


def _has_usable_signal(d: Dict) -> bool:
    """A row can only match an order if it carries phone, email, street+zip, or
    name+(zip|city). Name-only / city-only / zip-only rows are useless."""
    if norm_phone(d.get("phone")):
        return True
    if norm_email(d.get("email")):
        return True
    if norm_addr(d.get("street")) and norm_zip(d.get("zip")):
        return True
    if norm_text(d.get("full_name")) and (norm_zip(d.get("zip")) or norm_text(d.get("city"))):
        return True
    return False


def add_entry(data: Dict, created_by: str = "", source: str = "manual") -> Tuple[bool, str]:
    """Insert one entry. Returns (ok, message)."""
    clean = {f: (str(data.get(f)).strip() if data.get(f) not in (None, "") else None)
             for f in ("full_name", "phone", "email", "street", "city", "state", "zip", "reason")}
    if not _has_usable_signal(clean):
        return False, "至少要有 电话 / 邮箱 / (街道+邮编) / (姓名+邮编或城市) 其中一种,否则无法匹配任何订单"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = DBManager.get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """INSERT INTO order_system.customer_blacklist
                   (full_name, phone, email, street, city, state, zip, reason,
                    active, source, created_by, created_at)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,1,%s,%s,%s)""",
                (clean["full_name"], clean["phone"], clean["email"], clean["street"],
                 clean["city"], clean["state"], clean["zip"], clean["reason"],
                 source, created_by or None, now),
            )
        conn.commit()
        return True, "已添加"
    finally:
        conn.close()


def delete_entry(entry_id: int) -> None:
    conn = DBManager.get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                "DELETE FROM order_system.customer_blacklist WHERE id=%s", (entry_id,))
        conn.commit()
    finally:
        conn.close()


def _dup_key(d: Dict) -> Tuple:
    return (norm_phone(d.get("phone")), norm_email(d.get("email")),
            norm_addr(d.get("street")), norm_zip(d.get("zip")),
            norm_text(d.get("full_name")), norm_text(d.get("city")))


# =============================================================================
# Excel import + blank template
# =============================================================================

def make_blank_template() -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "黑名单"
    ws.append(BLACKLIST_COLUMNS)
    # one example row to show the shape (user deletes it)
    ws.append(["John Smith", "(415) 555-1234", "bad@example.com",
               "123 Main St Apt 4", "San Jose", "CA", "95112", "恶意退货"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    wb.close()
    return bio


def import_excel(file_storage, created_by: str = "") -> Dict[str, Any]:
    """Parse an uploaded .xlsx and bulk-add entries.
    Returns {imported, skipped, duplicate, skipped_rows:[(rownum,reason)]}.
    """
    wb = openpyxl.load_workbook(file_storage, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return {"imported": 0, "skipped": 0, "duplicate": 0,
                "skipped_rows": [], "error": "空文件"}
    # locate columns by header name (order-independent)
    col_idx = {}
    for i, h in enumerate(header):
        h = (str(h).strip() if h is not None else "")
        if h in _COL_TO_FIELD:
            col_idx[_COL_TO_FIELD[h]] = i
    if not col_idx:
        return {"imported": 0, "skipped": 0, "duplicate": 0, "skipped_rows": [],
                "error": f"表头不识别,需包含列: {', '.join(BLACKLIST_COLUMNS)}"}

    existing = {_dup_key(e) for e in list_entries(active_only=False)}
    imported = duplicate = skipped = 0
    skipped_rows: List[Tuple[int, str]] = []

    for rownum, raw in enumerate(rows, start=2):
        if raw is None or all(c in (None, "") for c in raw):
            continue
        d = {}
        for field, i in col_idx.items():
            val = raw[i] if i < len(raw) else None
            d[field] = (str(val).strip() if val not in (None, "") else None)
        if not _has_usable_signal(d):
            skipped += 1
            skipped_rows.append((rownum, "无可匹配信号(缺电话/邮箱/街道+邮编/姓名+邮编或城市)"))
            continue
        key = _dup_key(d)
        if key in existing:
            duplicate += 1
            continue
        existing.add(key)
        ok, _msg = add_entry(d, created_by=created_by, source="excel")
        if ok:
            imported += 1
        else:
            skipped += 1
            skipped_rows.append((rownum, _msg))
    wb.close()
    return {"imported": imported, "skipped": skipped, "duplicate": duplicate,
            "skipped_rows": skipped_rows}
