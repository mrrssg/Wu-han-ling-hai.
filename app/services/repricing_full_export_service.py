"""
Part 2: weekly full repricing export. Multi-store (macy_kuyotq + lowes_autool).

For every active offer of the given store:
  1. Calculate target origin_price using the latest supplier cost + Feishu
     config (returns 12% margin per the formula in repricing_formula).
  2. Compare against current DB origin_price.
  3. If different (delta >= $0.01), include the row in the output xlsx.
  4. If same, skip - no point uploading an unchanged price.

The output is a Mirakl-compatible `offers-import` xlsx that the operator
downloads and uploads to the Mirakl portal manually. We do NOT call OF24
from this path - this is the deliberate user choice.

Performance: this service pre-loads all required data in batch so the inner
loop is pure-Python with zero per-SKU DB roundtrips. 3000+ offers should
finish in well under a minute.

Audit trail: every iteration (skipped / would-update / blocked-by-alert)
writes a row to offer_price_change_log with run_type='full_export', so the
operator can review what was decided without opening the xlsx.
"""
import json
import os
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import shutil

from openpyxl import Workbook, load_workbook

from app.models.db_manager import DBManager
from app.services.repricing_formula import (
    calculate_breakdown,
    cost_from_supplier_price,
)
from app.services.repricing_monitor_service import (
    MAX_SUPPLIER_STALE_HOURS,
    get_supplier_freshness,
)
from app.services.repricing_stores import get_store


# Both stores are non-Dropship: a single `price` plus an optional discount
# (discount-price + discount window), NO retail-price/msrp. The actual column
# set is read from each store's base template at write time (write_xlsx) -
#   Macy  : 19 cols (offers_import_blank.xlsx)
#   Lowes : 18 cols, same minus `favorite-rank` (offers_import_lowes_blank.xlsx)
# _build_xlsx_row emits the 19-col superset; write_xlsx keeps only the columns
# the base template actually has. This constant is just the fallback header
# used when no base template file is found.
OFFERS_IMPORT_COLUMNS = [
    "sku", "product-id", "product-id-type", "description",
    "internal-description", "price", "price-additional-info", "quantity",
    "min-quantity-alert", "state", "available-start-date",
    "available-end-date", "logistic-class", "favorite-rank",
    "discount-start-date", "discount-end-date", "discount-price",
    "update-delete", "leadtime-to-ship",
]

MIN_PRICE_DELTA = 0.01    # ignore drift below 1¢


# =============================================================================
# Bulk preload helpers - replace per-SKU queries
# =============================================================================

def _load_active_offers(store_key: str) -> List[Dict]:
    scfg = get_store(store_key)
    conn = DBManager.get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """SELECT shop_sku, warehouse_sku, origin_price, raw_json,
                          state_code, quantity, last_cost_snapshot
                     FROM order_system.offerprice_listing
                    WHERE platform=%s AND shop_name=%s AND active=1""",
                (scfg["platform"], scfg["shop_name"]),
            )
            return cursor.fetchall() or []
    finally:
        conn.close()


def _load_pricing_configs(store_key: str) -> Dict[str, Dict]:
    conn = DBManager.get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """SELECT warehouse_sku, supplier, discount_factor, commission_rate,
                          return_shipping_base, length_in, width_in, height_in, weight_lb
                     FROM order_system.offer_pricing_config
                    WHERE store_key=%s""",
                (store_key,),
            )
            return {r["warehouse_sku"]: r for r in cursor.fetchall() or []}
    finally:
        conn.close()


def _load_supplier_prices() -> Dict[Tuple[str, str], Tuple[float, Any]]:
    """Return {(supplier, sku): (Price, Updated_At)} for both supplier tables."""
    out: Dict[Tuple[str, str], Tuple[float, Any]] = {}
    conn = DBManager.get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT SKU, Price, Updated_At FROM autooperate.newestdropship"
            )
            for r in cursor.fetchall() or []:
                sku = (r.get("SKU") or "").strip()
                if not sku:
                    continue
                p = r.get("Price")
                try:
                    p = float(p) if p is not None else None
                except (TypeError, ValueError):
                    p = None
                if p is not None:
                    out[("Costway", sku)] = (p, r.get("Updated_At"))

            cursor.execute(
                "SELECT SKU, Price, Updated_At FROM autooperate.newestdropship_vevor"
            )
            for r in cursor.fetchall() or []:
                sku = (r.get("SKU") or "").strip()
                if not sku:
                    continue
                p = r.get("Price")
                try:
                    p = float(p) if p is not None else None
                except (TypeError, ValueError):
                    p = None
                if p is not None:
                    out[("Vevor", sku)] = (p, r.get("Updated_At"))
    finally:
        conn.close()
    return out


def _load_blacklist(store_key: str) -> set:
    conn = DBManager.get_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """SELECT shop_sku FROM order_system.offer_alert_state
                    WHERE store_key=%s AND blacklisted=1""",
                (store_key,),
            )
            return {r["shop_sku"] for r in cursor.fetchall() or []}
    finally:
        conn.close()


# =============================================================================
# Per-row decision (pure-python, no DB)
# =============================================================================

def _decide(offer: Dict, cfg: Optional[Dict], sp_lookup: Dict, blacklist: set,
            formula_variant: str = "macy",
            discount_factor_override: Optional[float] = None,
            tier_row: Optional[Dict] = None) -> Dict[str, Any]:
    """Decide what to do with one offer; returns a result dict with one of
    these statuses:
      - 'skipped_delist_tier'      (定价方案判下架——不修价，等人工下架)
      - 'skipped_cold_watch'       (新品观察期——不动价)
      - 'skipped_blacklist'
      - 'alert_no_config'
      - 'alert_no_return_shipping'
      - 'alert_no_dim'
      - 'alert_no_supplier_price'
      - 'alert_unsupported_supplier'
      - 'skipped_same_price'        (no change needed)
      - 'would_update'              (include in xlsx)
    """
    shop_sku = offer["shop_sku"]
    warehouse_sku = offer.get("warehouse_sku")
    db_origin_price = offer.get("origin_price")

    # 分档定价联动（2026-07-17）：有档位的SKU按档位公式价导出；
    # 下架档不修价（等人工下架），新品观察档不动价
    tier = (tier_row or {}).get("tier")
    tier_target = (tier_row or {}).get("target_margin")
    if tier == "delist":
        return {"status": "skipped_delist_tier"}
    if tier == "cold_watch":
        return {"status": "skipped_cold_watch"}

    if shop_sku in blacklist:
        return {"status": "skipped_blacklist"}

    if not warehouse_sku:
        return {"status": "alert_no_config", "alert_type": "no_warehouse_sku"}

    if not cfg:
        return {"status": "alert_no_config", "alert_type": "feishu_config_missing"}

    supplier = cfg.get("supplier")
    if supplier not in ("Costway", "Vevor"):
        return {"status": "alert_unsupported_supplier", "alert_type": "unsupported_supplier"}

    rb = cfg.get("return_shipping_base")
    if rb is None:
        return {"status": "alert_no_return_shipping", "alert_type": "return_shipping_missing"}
    rb = float(rb)

    L = cfg.get("length_in"); W = cfg.get("width_in")
    H = cfg.get("height_in"); wt = cfg.get("weight_lb")
    if None in (L, W, H, wt) or any(float(v) == 0 for v in (L, W, H)):
        return {"status": "alert_no_dim", "alert_type": "dim_missing"}

    sp_info = sp_lookup.get((supplier, warehouse_sku))
    if not sp_info:
        return {"status": "alert_no_supplier_price", "alert_type": "cost_missing"}
    supplier_price, supplier_updated = sp_info

    new_cost = cost_from_supplier_price(supplier_price, supplier)
    # Store-level override (e.g. macy_kuyotq fixed at 0.4) wins over Feishu
    # per-SKU value to neutralise stale/dirty Feishu data.
    if discount_factor_override is not None:
        df = float(discount_factor_override)
    else:
        feishu_df = cfg.get("discount_factor")
        if feishu_df is None:
            return {"status": "alert_no_config", "alert_type": "discount_factor_missing"}
        df = float(feishu_df)
    cr = float(cfg["commission_rate"]) if cfg.get("commission_rate") is not None else 0.0

    # 档位除数：1 − 佣金 − 档位目标毛利（12/15/18档各不同）；无档位记录用公式默认
    divisor_override = None
    if tier_target is not None and cfg.get("commission_rate") is not None:
        divisor_override = 1.0 - cr - float(tier_target)

    bd = calculate_breakdown(
        supplier=supplier,
        supplier_price=supplier_price,
        return_shipping_base=rb,
        discount_factor=df,
        length_in=float(L), width_in=float(W),
        height_in=float(H), weight_lb=float(wt),
        formula_variant=formula_variant,
        divisor_override=divisor_override,
    )
    target_origin = round(float(bd.origin_price), 2)
    target_discount = round(float(bd.discount_price), 2)

    if db_origin_price is None:
        # treat as needs-update so the operator can fill it in
        delta = None
    else:
        delta = abs(target_origin - float(db_origin_price))

    if delta is not None and delta < MIN_PRICE_DELTA:
        return {
            "status": "skipped_same_price",
            "supplier": supplier,
            "supplier_price": supplier_price,
            "supplier_updated": supplier_updated,
            "new_cost": new_cost,
            "target_origin_price": target_origin,
            "current_origin_price": float(db_origin_price),
            "delta": delta,
        }

    return {
        "status": "would_update",
        "supplier": supplier,
        "supplier_price": supplier_price,
        "supplier_updated": supplier_updated,
        "new_cost": new_cost,
        "target_origin_price": target_origin,
        "target_discount_price": target_discount,
        "current_origin_price": float(db_origin_price) if db_origin_price is not None else None,
        "delta": delta,
        "return_shipping_extra": bd.return_shipping_extra,
        "return_cost_estimate": bd.return_cost_estimate,
        "total_cost": bd.total_cost,
        "formula_calc_price": bd.formula_calc_price,
        "discount_factor": df,
        "commission_rate": cr,
        "return_shipping_base": rb,
    }


# =============================================================================
# 生成后逐行复验（2026-07-17用户拍板：有一行不过就不给下载，fail-closed）
# =============================================================================

def _verify_xlsx(path: str, intents: Dict[str, Dict], push_discount: bool) -> List[str]:
    """把写完的文件重新读回来，和算价意图逐行核对（专防写文件环节的错列/错行/漏行）：
    ① SKU集合与意图完全一致（不多行/不缺行/不重复）
    ② 原价、折扣价与意图一致
    ③ 原价 = 折扣价 ÷ 折扣系数
    ④ 保本线：折扣价×(1−佣金) ≥ 成本
    ⑤ 折扣结束日期未过期（过期=折扣不生效，买家看到原价）
    返回失败清单，空=全部通过。"""
    fails: List[str] = []
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    head = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(head)}
    seen = set()
    today = datetime.now().date()
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = row[idx["sku"]]
        if sku is None:
            continue
        it = intents.get(sku)
        if not it:
            fails.append(f"{sku}: 文件里多出的行（不在算价意图中）")
            continue
        if sku in seen:
            fails.append(f"{sku}: 重复行")
            continue
        seen.add(sku)
        try:
            price = float(row[idx["price"]])
        except (TypeError, ValueError):
            fails.append(f"{sku}: price列不是数字")
            continue
        if abs(price - it["origin"]) > 0.011:
            fails.append(f"{sku}: 原价{price}≠意图{it['origin']}")
        if push_discount:
            dcell = row[idx["discount-price"]]
            disc = float(dcell) if dcell not in (None, "") else None
            if disc is None:
                fails.append(f"{sku}: 缺折扣价")
            else:
                if it.get("discount") is not None and abs(disc - it["discount"]) > 0.011:
                    fails.append(f"{sku}: 折扣价{disc}≠意图{it['discount']}")
                if it.get("df") and abs(price - round(disc / float(it["df"]), 2)) > 0.011:
                    fails.append(f"{sku}: 原价≠折扣÷{it['df']}")
                cr = float(it.get("cr") or 0)
                if it.get("cost") is not None and disc * (1 - cr) < float(it["cost"]) - 0.01:
                    fails.append(f"{sku}: 破保本线（折扣{disc}×{1-cr:.2f} < 成本{float(it['cost']):.2f}）")
            dend = row[idx["discount-end-date"]]
            if dend:
                try:
                    if datetime.strptime(str(dend)[:10], "%Y-%m-%d").date() < today:
                        fails.append(f"{sku}: 折扣窗口已过期({dend})——上传后折扣不生效，买家看到原价")
                except ValueError:
                    pass
    missing = set(intents) - seen
    if missing:
        fails.append(f"文件缺{len(missing)}行，如: {sorted(missing)[:5]}")
    return fails


# =============================================================================
# Excel writer
# =============================================================================

def _raw_logistic_code(raw: Dict):
    lc = raw.get("logistic_class")
    if isinstance(lc, dict):
        return lc.get("code")
    return lc


def _raw_first(raw: Dict, key: str):
    """raw['prices'] / raw['retail_prices'] is a list; return first dict or {}."""
    v = raw.get(key)
    if isinstance(v, list) and v and isinstance(v[0], dict):
        return v[0]
    return {}


def _build_xlsx_row(offer: Dict, decision: Dict, raw: Dict,
                    push_discount: bool) -> Dict[str, Any]:
    """Build one offers-import xlsx row (standard non-Dropship 19-col layout).

    Every store writes the customer price into `price` (= 活动前原价).
    push_discount stores (Lowes) also write the 折扣后价格 into `discount-price`
    and reuse the live offer's discount window (user rule: 日期沿用现有的).
    Price-only stores (Macy) leave the discount columns blank.
    """
    row = {
        "sku": offer["shop_sku"],
        "product-id": raw.get("product_sku") or offer["shop_sku"],
        "product-id-type": "SKU",
        "description": None,
        "internal-description": None,
        "price": decision["target_origin_price"],
        "price-additional-info": None,
        "quantity": raw.get("quantity") or offer.get("quantity") or 0,
        "min-quantity-alert": None,
        "state": raw.get("state_code") or offer.get("state_code") or "11",
        "available-start-date": None,
        "available-end-date": None,
        "logistic-class": _raw_logistic_code(raw),
        "favorite-rank": None,
        "discount-start-date": None,
        "discount-end-date": None,
        "discount-price": None,
        "update-delete": "update",
        "leadtime-to-ship": raw.get("leadtime_to_ship"),
    }

    if push_discount:
        prices0 = _raw_first(raw, "prices")
        row["discount-price"] = decision.get("target_discount_price")
        row["discount-start-date"] = prices0.get("discount_start_date")
        row["discount-end-date"] = prices0.get("discount_end_date")

    return row


def write_xlsx(rows: List[Dict[str, Any]], output_path: str,
                base_template_path: Optional[str] = None) -> int:
    """Write the offers-import xlsx.

    If `base_template_path` exists, we copy it and append data rows to
    preserve header styling, freeze panes, column widths, etc. Otherwise
    we fall back to creating a plain Workbook from scratch.
    """
    if base_template_path and os.path.exists(base_template_path):
        # Copy base then append data rows after header
        shutil.copyfile(base_template_path, output_path)
        wb = load_workbook(output_path)
        sheet_name = "offers-import" if "offers-import" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        # Clear any leftover data rows (defensive: base should already be header-only)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        # Discover header order from the base (lower-case match for safety)
        base_headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        # Build header-name -> column-index map (1-based)
        col_idx_by_name = {h: i + 1 for i, h in enumerate(base_headers) if h}
        # Append rows aligned to base header order
        for r in rows:
            line = []
            for h in base_headers:
                line.append(r.get(h))
            ws.append(line)
        wb.save(output_path)
        wb.close()
        return len(rows)

    # Fallback: plain workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "offers-import"
    ws.append(OFFERS_IMPORT_COLUMNS)
    for r in rows:
        ws.append([r.get(c) for c in OFFERS_IMPORT_COLUMNS])
    wb.save(output_path)
    return len(rows)


# =============================================================================
# Audit log
# =============================================================================

def _log_decisions(run_id: str, decisions: List[Tuple[Dict, Dict]], store_key: str):
    """Bulk-insert decision rows into offer_price_change_log."""
    if not decisions:
        return
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    rows = []
    for offer, d in decisions:
        status = d["status"]
        status_for_log = {
            "skipped_blacklist": "skipped",
            "alert_no_config": "alert",
            "alert_no_return_shipping": "alert",
            "alert_no_dim": "alert",
            "alert_no_supplier_price": "alert",
            "alert_unsupported_supplier": "alert",
            "skipped_same_price": "skipped",
            "would_update": "dry_run",   # not pushed yet; "dry_run" matches existing taxonomy
        }.get(status, "skipped")

        rows.append((
            run_id,
            "full_export",
            store_key,
            offer["shop_sku"],
            offer.get("warehouse_sku"),
            now,
            status_for_log,
            (
                "calc==db, no update needed" if status == "skipped_same_price"
                else "blacklisted" if status == "skipped_blacklist"
                else f"calc={d.get('target_origin_price')} vs db={d.get('current_origin_price')}, delta={d.get('delta')}" if status == "would_update"
                else status
            ),
            d.get("alert_type"),
            d.get("supplier"),
            d.get("supplier_price"),
            d.get("new_cost"),
            d.get("current_origin_price"),
            d.get("target_origin_price"),
            d.get("target_discount_price"),
            d.get("discount_factor"),
            d.get("commission_rate"),
            d.get("return_shipping_base"),
            d.get("return_shipping_extra"),
            d.get("return_cost_estimate"),
            d.get("total_cost"),
            d.get("formula_calc_price"),
            d.get("target_origin_price"),
        ))

    sql = """
        INSERT INTO order_system.offer_price_change_log
            (run_id, run_type, store_key, shop_sku, warehouse_sku, triggered_at,
             status, decision_reason, alert_type,
             supplier, supplier_price_db, new_cost,
             old_origin_price, new_origin_price, new_discount_price,
             discount_factor, commission_rate, return_shipping_base,
             return_shipping_extra, return_cost_estimate, total_cost,
             formula_calc_price, target_origin_price)
        VALUES (%s, %s, %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s)
    """
    conn = DBManager.get_connection()
    try:
        with conn.cursor() as cursor:
            chunk = 500
            for i in range(0, len(rows), chunk):
                cursor.executemany(sql, rows[i:i + chunk])
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# =============================================================================
# Top-level entry
# =============================================================================

def run_full_export(output_dir: str, store_key: str = "macy_kuyotq") -> Dict[str, Any]:
    """Top-level entry. Returns a summary dict including output_file path.
    Aborts if supplier data is stale.
    """
    scfg = get_store(store_key)               # raises if unsupported
    if scfg.get("offer_sync_only"):
        return {
            "success": False,
            "store_key": store_key,
            "msg": (f"store {store_key} is offer_sync_only - full export requires "
                    f"Feishu pricing config; provision it first"),
        }
    push_discount = scfg["push_discount"]
    formula_variant = scfg["formula_variant"]
    discount_factor_override = scfg.get("discount_factor_override")

    started = datetime.now()
    run_id = f"full-{store_key}-{started.strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6]}"

    freshness = get_supplier_freshness()
    if freshness["costway_stale"] or freshness["vevor_stale"]:
        return {
            "success": False,
            "run_id": run_id,
            "store_key": store_key,
            "msg": "supplier data stale; refusing to export",
            "freshness": {
                "costway_max": str(freshness["costway_max"]),
                "vevor_max": str(freshness["vevor_max"]),
                "threshold_hours": freshness["threshold_hours"],
            },
        }

    print(f"[{run_id}] preloading all data ...")
    active_offers = _load_active_offers(store_key)
    configs = _load_pricing_configs(store_key)
    sp_lookup = _load_supplier_prices()
    blacklist = _load_blacklist(store_key)
    # 分档定价（2026-07-17联动）：有档位的SKU导出价=档位公式价
    tier_map: Dict[str, Dict] = {}
    try:
        conn = DBManager.get_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute(
                    """SELECT shop_sku, tier, target_margin
                         FROM order_system.pricing_tier WHERE store_key=%s""",
                    (store_key,))
                tier_map = {r["shop_sku"]: r for r in cursor.fetchall() or []}
        finally:
            conn.close()
    except Exception:
        tier_map = {}
    print(f"  active offers : {len(active_offers)}")
    print(f"  configs       : {len(configs)}")
    print(f"  supplier prices: {len(sp_lookup)}")
    print(f"  blacklist     : {len(blacklist)}")
    print(f"  pricing tiers : {len(tier_map)}")

    decisions: List[Tuple[Dict, Dict]] = []
    xlsx_rows: List[Dict[str, Any]] = []

    summary = {
        "total_offers": len(active_offers),
        "would_update": 0,
        "skipped_same_price": 0,
        "skipped_blacklist": 0,
        "alert_no_config": 0,
        "alert_no_return_shipping": 0,
        "alert_no_dim": 0,
        "alert_no_supplier_price": 0,
        "alert_unsupported_supplier": 0,
    }

    from app.services.pricing_plan_service import COLD_BATCH

    by_tier: Dict[str, int] = {}
    for offer in active_offers:
        wh = offer.get("warehouse_sku")
        cfg = configs.get(wh) if wh else None
        tier_row = tier_map.get(offer["shop_sku"])
        decision = _decide(offer, cfg, sp_lookup, blacklist,
                           formula_variant=formula_variant,
                           discount_factor_override=discount_factor_override,
                           tier_row=tier_row)
        # 零销量促活与API管道共用分批阀：每轮最多COLD_BATCH个（用户2026-07-17定：导出也压批）
        # 这批传完夜间同步后价格变成"已一致"，下一轮自动放下一批
        if decision["status"] == "would_update" \
                and (tier_row or {}).get("tier") == "cold_12" \
                and by_tier.get("cold_12", 0) >= COLD_BATCH:
            decision = {"status": "skipped_cold_deferred"}
        decisions.append((offer, decision))
        summary[decision["status"]] = summary.get(decision["status"], 0) + 1

        if decision["status"] == "would_update":
            tname = (tier_row or {}).get("tier") or "(无档位)"
            by_tier[tname] = by_tier.get(tname, 0) + 1
            raw = {}
            if offer.get("raw_json"):
                try:
                    raw = json.loads(offer["raw_json"])
                except (TypeError, ValueError):
                    raw = {}
            xlsx_rows.append(_build_xlsx_row(offer, decision, raw, push_discount))

    # Sort xlsx output by sku for tidy review
    xlsx_rows.sort(key=lambda r: r["sku"])

    os.makedirs(output_dir, exist_ok=True)
    fname = f"{store_key}_repricing_{started.strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(output_dir, fname)

    # Look for the store's styled base template under instance/repricing/
    base_template = os.path.join(
        os.path.dirname(os.path.dirname(output_dir)),  # instance/
        "repricing",
        scfg["excel_template"],
    )

    written = write_xlsx(xlsx_rows, output_path, base_template_path=base_template)

    # 生成后逐行复验（fail-closed）：任何一行不过 → 文件隔离，不提供下载
    intents: Dict[str, Dict] = {}
    for offer, d in decisions:
        if d["status"] == "would_update":
            intents[offer["shop_sku"]] = {
                "origin": float(d["target_origin_price"]),
                "discount": (float(d["target_discount_price"])
                             if d.get("target_discount_price") is not None else None),
                "cost": d.get("new_cost"),
                "cr": d.get("commission_rate"),
                "df": d.get("discount_factor"),
            }
    verify_fails = _verify_xlsx(output_path, intents, push_discount)
    _log_decisions(run_id, decisions, store_key)
    if verify_fails:
        rejected = output_path + ".REJECTED"
        os.replace(output_path, rejected)
        print(f"[{run_id}] VERIFY FAILED ({len(verify_fails)}): {verify_fails[:5]}")
        summary["by_tier"] = by_tier
        return {
            "success": False,
            "run_id": run_id,
            "store_key": store_key,
            "msg": f"生成后复验不通过（{len(verify_fails)}项）——文件已隔离，不提供下载",
            "verify_failures": verify_fails[:20],
            "rejected_file": rejected,
            "summary": summary,
        }
    summary["verify"] = {"passed": True, "rows_checked": written}

    # NB: Feishu supplier-price writeback used to happen here, but it was
    # premature - this endpoint only GENERATES the xlsx; Mirakl prices change
    # only after the operator manually uploads it. Writing back at this point
    # made Feishu Formula columns show the new state before Mirakl actually
    # had it, plus the ~20s Feishu full-table scan made the HTTP response
    # exceed client timeout (499 incidents). Writeback now happens only in
    # push_one / push_batch where OF24 was actually called.

    duration = (datetime.now() - started).total_seconds()
    summary["by_tier"] = by_tier
    return {
        "success": True,
        "run_id": run_id,
        "store_key": store_key,
        "output_file": output_path,
        "filename": fname,
        "rows_written": written,
        "summary": summary,
        "duration_seconds": round(duration, 2),
        "freshness": {
            "costway_max": str(freshness["costway_max"]),
            "vevor_max": str(freshness["vevor_max"]),
        },
    }
