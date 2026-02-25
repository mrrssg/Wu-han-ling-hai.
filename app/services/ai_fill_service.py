import json
import os
import threading
import uuid
from datetime import datetime
from typing import Dict, List, Tuple

import pandas as pd
import openpyxl
from openai import OpenAI


MODEL_NAME = "gpt-5.1"

SHEET_DATA = "Data"
SHEET_COLS = "Columns"
SHEET_REF = "ReferenceData"

DATA_HEADER_ROW = 1
DATA_START_ROW = 3

HEADER_ROW = 1
ANCHOR_HEADER = "Product Category"
START_FROM_ROW = 3

IDX_LABEL = 2
IDX_DESC = 3
IDX_REQ = 5
START_ROW_COLS = 2

MAX_OPTIONS_THRESHOLD_DEFAULT = 60

DEFAULT_WHITELIST: List[str] = []

DEFAULT_FIXED_VALUES: Dict[str, str] = {
    "Is this product sold exclusively to and by The Home Depot?": "No",
    "Is this a new version of an existing item?": "No",
    "COUNTRY OF ORIGIN": "CN",
    "Country of Origin Name": "CHINA",
    "Sell UOM (as sold to consumer)": "EA-Each",
    "Made-To-Order": "No",
    "Does the item contain Mercury (ex: fluorescent light bulb, HVAC, switch, thermostat)?": "N",
    "Is the item a liquid or contain a liquid (this does not include appliances or heaters that contain totally enclosed liquids)?": "N",
    "Is the item a chemical / solvent or contain a chemical / solvent?": "N",
    "Is the item an aerosol or contain an aerosol?": "N",
    "Is the item a pesticide or contain a pesticide, herbicide, fungicide?": "N",
    "Is the item or does the item contain a battery (lithium, alkaline, lead-acid, etc.)?": "N",
    "Is the item or does the item contain a compressed gas?": "N",
    "Sellable Unit?": "Y",
    "Are your products labeled with age - grading or otherwise packaged, labeled or marketed for children?": "No",
    "Is your product intended to be put into childrens mouths, intended to be applied to childrens bodies, or is it mouthable (able to be sucked or chewed) by children under 3 years of age?": "No",
    "Is your product primarily designed and intended for children 12 years of age and under?": "No",
    "Will children be exposed to your product for more than an hour (Ex. clothing, footwear, jewelry, certain toys)?": "No",
    "Is this product regulated by a type of VOC guideline or rule at the state level?": "No",
    "Eco Actions": "No",
    "ExcludedShipStates": "Guam;Hawaii;Alaska;Virgin Islands;Puerto Rico",
    "Proposition 65 warning required?": "No",
    "IBI Lithium Battery Flag": "N",
    "Can this item be shipped anywhere in the US?": "No",
    "Sell Pkg Qty (as sold to consumer)": "1",
    "Vendor Processing Days": "2",
    "Ship From YOW/CHUB": "CHUB_OMS",
    "Does the textile in your product contain one or more PFAS chemicals in any amount?": "No",
    "Is your product considered outdoor apparel for severe wet weather conditions, as the California Safer Clothing and Textiles Act defines that term?": "No",
    "Do you have the Certificate of Compliance - PFAS in Textiles?": "No",
}


def get_instance_dir(base_dir: str) -> str:
    return os.path.join(base_dir, "instance", "ai_fill")


def get_config_path(base_dir: str) -> str:
    return os.path.join(get_instance_dir(base_dir), "config.json")


def load_config(base_dir: str) -> Dict:
    cfg = {
        "whitelist": list(DEFAULT_WHITELIST),
        "fixed_values": dict(DEFAULT_FIXED_VALUES),
        "max_options_threshold": MAX_OPTIONS_THRESHOLD_DEFAULT,
    }
    path = get_config_path(base_dir)
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                cfg.update({k: data[k] for k in data.keys() if k in cfg})
        except Exception:
            pass
    return cfg


def save_config(base_dir: str, cfg: Dict) -> None:
    os.makedirs(get_instance_dir(base_dir), exist_ok=True)
    path = get_config_path(base_dir)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def _write_progress(path: str, data: Dict) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp_path = path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    os.replace(tmp_path, path)


def _read_progress(path: str) -> Dict:
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _now_iso() -> str:
    return datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")


def fill_fixed_values(ws_data, fixed_values: Dict[str, str]) -> int:
    col_map = {}
    anchor_col_idx = None
    for cell in ws_data[HEADER_ROW]:
        if cell.value:
            header = str(cell.value).strip()
            col_map[header] = cell.column
            if header == ANCHOR_HEADER:
                anchor_col_idx = cell.column

    if anchor_col_idx is None:
        return 0

    filled = 0
    max_row = ws_data.max_row
    for row_idx in range(START_FROM_ROW, max_row + 1):
        anchor_val = ws_data.cell(row=row_idx, column=anchor_col_idx).value
        if anchor_val and str(anchor_val).strip():
            for header_name, val in fixed_values.items():
                if header_name in col_map:
                    col_idx = col_map[header_name]
                    ws_data.cell(row=row_idx, column=col_idx).value = val
            filled += 1
    return filled


def load_column_rules(wb, whitelist: List[str]) -> Dict[str, Dict]:
    ws = wb[SHEET_COLS]
    rules = {}

    for row in ws.iter_rows(min_row=START_ROW_COLS, values_only=True):
        code = row[0]
        label = row[IDX_LABEL - 1]
        desc = row[IDX_DESC - 1]
        req = row[IDX_REQ - 1]

        if not label:
            continue
        label_str = str(label).strip()
        is_req = "REQUIRED" in str(req).upper() if req else False
        is_in_whitelist = label_str in whitelist

        if is_req or is_in_whitelist:
            rules[label_str] = {
                "code": str(code).strip() if code else "",
                "required": is_req,
                "desc": str(desc).strip() if desc else "No description",
            }

    return rules


def load_ref_options(wb, rules: Dict[str, Dict]) -> Dict[str, List[str]]:
    ws = wb[SHEET_REF]
    ref_data = {}

    valid_code_map = {}
    for label, info in rules.items():
        if info.get("code"):
            valid_code_map[str(info["code"]).strip()] = label

    for col in ws.iter_cols(min_row=1, values_only=True):
        code_in_ref = col[0]
        if not code_in_ref:
            continue
        code_str = str(code_in_ref).strip()
        if code_str not in valid_code_map:
            continue
        target_label = valid_code_map[code_str]
        options = [str(x).strip() for x in col[2:] if x]
        if options:
            ref_data[target_label] = options

    return ref_data


def _build_tasks(rules, ref_options, data_header_map, whitelist, max_options_threshold, ws_data, current_hd_row):
    tasks = {}
    for field_name, rule_info in rules.items():
        if field_name not in data_header_map:
            continue
        if not (rule_info["required"] or field_name in whitelist):
            continue

        col_idx = data_header_map[field_name]
        current_val = ws_data.cell(row=current_hd_row, column=col_idx).value
        if current_val and str(current_val).strip():
            continue

        valid_opts = ref_options.get(field_name, [])
        if valid_opts:
            if len(valid_opts) > max_options_threshold:
                option_instruction = (
                    "Standard Field. Please use the most professional standard term. "
                    "Do not invent unusual words."
                )
            else:
                option_instruction = f"MUST CHOOSE ONE FROM: {json.dumps(valid_opts, ensure_ascii=False)}"
        else:
            option_instruction = "Free text input."

        tasks[field_name] = {
            "Description": rule_info["desc"],
            "Constraint": option_instruction,
        }

    return tasks


def _call_openai(row_src, tasks, img_url):
    source_info = {
        "Title": row_src.get("Item Name"),
        "Spec": row_src.get("Specification"),
        "Desc": row_src.get("Description"),
        "Source_SKU": row_src.get("SKU"),
    }

    prompt = (
        "You are a Home Depot data expert. Fill the fields based on source data.\n\n"
        "[Source Data]\n"
        f"{json.dumps(source_info, ensure_ascii=False)}\n\n"
        "[Field Requirements]\n"
        f"{json.dumps(tasks, ensure_ascii=False, indent=2)}\n\n"
        "[Must Do]\n"
        "1) Every field must return a value. Never return null or empty string.\n"
        "2) If missing, infer from image; if unknown, use the most common value for that category.\n"
        "3) Dropdown values must match exactly. (NO multi-select: return ONLY ONE value, no commas or slashes.)\n"
    )

    client = OpenAI()

    messages = [
        {"role": "system", "content": "Output valid JSON only."},
        {"role": "user", "content": [{"type": "text", "text": prompt}]},
    ]

    if img_url and str(img_url).startswith("http"):
        messages[1]["content"].append({"type": "image_url", "image_url": {"url": img_url}})

    response = client.chat.completions.create(
        model=MODEL_NAME,
        response_format={"type": "json_object"},
        messages=messages,
    )

    raw_content = response.choices[0].message.content
    if raw_content is None:
        raise ValueError("OpenAI response content is empty")

    return json.loads(raw_content)


def start_job(
    base_dir: str,
    source_path: str,
    template_path: str,
    output_path: str,
    config: Dict,
    progress_path: str,
    job_id: str | None = None,
) -> str:
    if not job_id:
        job_id = str(uuid.uuid4())

    init_progress = {
        "job_id": job_id,
        "status": "running",
        "message": "starting",
        "started_at": _now_iso(),
        "updated_at": _now_iso(),
        "total": 0,
        "current": 0,
        "current_sku": "",
        "rows": [],
        "output_file": "",
        "error": "",
    }
    _write_progress(progress_path, init_progress)

    thread = threading.Thread(
        target=_run_job,
        args=(base_dir, source_path, template_path, output_path, config, progress_path),
        daemon=True,
    )
    thread.start()
    return job_id


def _run_job(base_dir: str, source_path: str, template_path: str, output_path: str, config: Dict, progress_path: str):
    progress = _read_progress(progress_path)
    try:
        if not os.environ.get("OPENAI_API_KEY"):
            key_path = os.path.join(base_dir, "instance", "openai_key.txt")
            if os.path.exists(key_path):
                with open(key_path, "r", encoding="utf-8") as f:
                    key = f.read().strip()
                if key:
                    os.environ["OPENAI_API_KEY"] = key

        if not os.environ.get("OPENAI_API_KEY"):
            raise RuntimeError("OPENAI_API_KEY is not set")

        df_source = pd.read_excel(source_path).fillna("")
        wb = openpyxl.load_workbook(template_path)
        ws_data = wb[SHEET_DATA]

        fill_fixed_values(ws_data, config["fixed_values"])

        rules = load_column_rules(wb, config["whitelist"])
        ref_options = load_ref_options(wb, rules)

        data_header_map = {}
        for cell in ws_data[DATA_HEADER_ROW]:
            if cell.value:
                data_header_map[str(cell.value).strip()] = cell.column

        total = len(df_source)
        progress["total"] = total
        progress["message"] = "running"
        progress["updated_at"] = _now_iso()
        _write_progress(progress_path, progress)

        for i in range(total):
            row_src = df_source.iloc[i]
            current_hd_row = DATA_START_ROW + i
            sku = row_src.get("SKU", f"Row {i + 1}")

            tasks = _build_tasks(
                rules=rules,
                ref_options=ref_options,
                data_header_map=data_header_map,
                whitelist=config["whitelist"],
                max_options_threshold=config["max_options_threshold"],
                ws_data=ws_data,
                current_hd_row=current_hd_row,
            )

            row_log = {
                "row": i + 1,
                "sheet_row": current_hd_row,
                "sku": str(sku),
                "status": "pending",
                "filled": {},
                "error": "",
            }

            if not tasks:
                row_log["status"] = "skipped"
            else:
                try:
                    img_url = row_src.get("Images1")
                    ai_data = _call_openai(row_src, tasks, img_url)
                    for key, val in ai_data.items():
                        if key in data_header_map and val is not None and str(val).strip() != "":
                            col_idx = data_header_map[key]
                            ws_data.cell(row=current_hd_row, column=col_idx).value = val
                            row_log["filled"][key] = val
                    row_log["status"] = "filled"
                except Exception as e:
                    row_log["status"] = "error"
                    row_log["error"] = str(e)

            progress["current"] = i + 1
            progress["current_sku"] = str(sku)
            progress["rows"].append(row_log)
            progress["updated_at"] = _now_iso()
            _write_progress(progress_path, progress)

        wb.save(output_path)

        progress["status"] = "done"
        progress["message"] = "completed"
        progress["output_file"] = output_path
        progress["updated_at"] = _now_iso()
        _write_progress(progress_path, progress)

    except Exception as e:
        progress["status"] = "error"
        progress["error"] = str(e)
        progress["updated_at"] = _now_iso()
        _write_progress(progress_path, progress)
