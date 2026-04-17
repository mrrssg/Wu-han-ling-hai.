import os

import io

import requests

import zipfile

import pandas as pd

from io import StringIO

from datetime import datetime

from typing import Iterable, Optional, Tuple

from app.models.db_manager import DBManager  # 引入我们的数据库管理器

from urllib3.util.retry import Retry

import urllib3

import time      # <--- 主要是漏了这个！

import random    # <--- GIGA 签名需要

import string    # <--- GIGA 签名需要

import hmac      # <--- GIGA 签名需要

import hashlib   # <--- GIGA 签名需要

import base64    # <--- GIGA 签名需要

from openpyxl import load_workbook

from requests.adapters import HTTPAdapter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)





class StockService:

    # 定义下载链接

    URL_COSTWAY = "https://cdn.costway.com/media/feed/dropship.zip"

    URL_VEVOR = "https://ads-feed.s3.us-west-2.amazonaws.com/ads/business/553/vevor-553.xlsx"

    URL_SONGMICS = "https://emes-us.songmics.com/us/songmics-product-2b.csv"

    HD_BASE_URL = "https://api.teapplix.com/api2/ProductQuantity"

    HD_WAREHOUSE_ID = 3

    HD_POST_TYPE = "in-stock"

    GIGA_CLIENT_ID = "39f40814-da69-4b2e-9acd-4e9969c6380c"

    GIGA_CLIENT_SECRET = "6dda9239fc0d4bc1a8a0da426c1aece9"

    GIGA_BASE_URL = "https://openapi.gigab2b.com"  # 生产环境



    @staticmethod

    def sync_all_suppliers():

        """

        主入口：同步所有供应商

        返回: (bool, string) -> (是否全部成功, 详细消息)

        """

        msgs = []

        status_c = False

        status_v = False



        # load costway zip password from config
        try:
            from flask import current_app
            base_dir = current_app.config.get('BASE_DIR', os.getcwd())
        except Exception:
            base_dir = os.getcwd()
        _cfg = StockService.load_hd_config(base_dir)
        _costway_pwd = _cfg.get("costway_zip_password", "")
        _costway_passwords = [_costway_pwd] if _costway_pwd else None

        # 1. 更新 Costway

        try:

            t0 = time.time()

            print("==[1] COSTWAY start==")

            csv_text = StockService.download_csv(StockService.URL_COSTWAY, passwords=_costway_passwords)

            print(f"==[1] COSTWAY download done, {time.time() - t0:.2f}s ==")

            if csv_text:

                t1 = time.time()

                success, msg = StockService.process_costway_data(csv_text)

                print(f"==[1] COSTWAY db done, {time.time() - t1:.2f}s ==")

                status_c = success

                msgs.append(f"Costway: {msg}")

            else:

                msgs.append("Costway: 下载失败或解压为空")

        except Exception as e:

            msgs.append(f"Costway异常: {str(e)}")



        # 2. 更新 Vevor

        try:

            t2 = time.time()

            print("==[2] VEVOR start==")

            xlsx_bytes = StockService.download_xlsx(StockService.URL_VEVOR)

            print(f"==[2] VEVOR download done, {time.time() - t2:.2f}s ==")

            if xlsx_bytes:

                t3 = time.time()

                success, msg = StockService.process_vevor_data(xlsx_bytes)

                print(f"==[2] VEVOR db done, {time.time() - t3:.2f}s ==")

                status_v = success

                msgs.append(f"Vevor: {msg}")

            else:

                msgs.append("Vevor: 下载失败")

        except Exception as e:

            msgs.append(f"Vevor异常: {str(e)}")



        # 3. Songmics

        try:

            t3 = time.time()

            print("==[3] SONGMICS start==")

            csv_text = StockService.download_csv(StockService.URL_SONGMICS)

            print(f"==[3] SONGMICS download done, {time.time() - t3:.2f}s ==")

            if csv_text:

                t4 = time.time()

                success, msg = StockService.process_songmics_data(csv_text)

                print(f"==[3] SONGMICS db done, {time.time() - t4:.2f}s ==")

                msgs.append(f"Songmics: {msg}")

            else:

                msgs.append("Songmics: download failed")

        except Exception as e:

            msgs.append(f"Songmics error: {str(e)}")



        try:

            t4 = time.time()

            print("==[4] GIGA start==")

            # 调用我们在下面写的 sync_giga_stock 方法

            success, msg = StockService.sync_giga_stock()

            print(f"==[4] GIGA done, {time.time() - t4:.2f}s ==")

            msgs.append(f"GIGA: {msg}")

        except Exception as e:

            msgs.append(f"GIGA异常: {str(e)}")





        # 汇总结果

        final_msg = " | ".join(msgs)

        return (status_c and status_v), final_msg







    # =================================================

    #  以下是你提供的逻辑代码 (稍作修改以适配类结构)

    # =================================================



    @staticmethod

    def download_csv(url: str, passwords: Optional[Iterable[str]] = None) -> Optional[str]:

        default_passwords = ["915473", "980178", "589726", "679215", "478329"]

        pwd_list = list(passwords) if passwords else default_passwords



        headers = {

            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

        }



        # ✅ 强化：重试 + 退避 + 断点续传下载zip（防SSL EOF/网络抖动）

        session = requests.Session()

        retry_strategy = Retry(

            total=8,

            connect=8,

            read=8,

            backoff_factor=2,  # 2,4,8,16...

            status_forcelist=[429, 500, 502, 503, 504],

            allowed_methods=["GET"]

        )

        adapter = HTTPAdapter(max_retries=retry_strategy, pool_connections=100, pool_maxsize=100)

        session.mount("https://", adapter)

        session.mount("http://", adapter)



        try:

            print(f"正在下载：{url}")



            # ✅ 先用 HEAD 判断是否zip / content-length

            is_zip_url = url.lower().endswith(".zip")



            # ✅ 如果是zip，建议落盘下载，避免resp.content一次性读完导致EOF/内存问题

            if is_zip_url:

                save_path = os.path.join(os.getcwd(), "temp_download.zip")

                temp_path = save_path + ".part"



                # 断点续传

                resume = 0

                if os.path.exists(temp_path):

                    resume = os.path.getsize(temp_path)



                dl_headers = dict(headers)

                if resume > 0:

                    dl_headers["Range"] = f"bytes={resume}-"



                for attempt in range(1, 9):

                    try:

                        with session.get(url, headers=dl_headers, stream=True, timeout=60, verify=False) as r:

                            r.raise_for_status()



                            mode = "ab" if "Range" in dl_headers else "wb"

                            with open(temp_path, mode) as f:

                                for chunk in r.iter_content(chunk_size=1024 * 1024):

                                    if chunk:

                                        f.write(chunk)



                        os.replace(temp_path, save_path)



                        with open(save_path, "rb") as f:

                            content = f.read()



                        break

                    except Exception as e:

                        print(f"下载失败(第{attempt}次)：{e}")

                        time.sleep(2 * attempt)

                else:

                    return None



            else:

                resp = session.get(url, headers=headers, timeout=60, verify=False)

                resp.raise_for_status()

                content = resp.content



            # ✅ 判断zip

            is_zip = content[:4] == b'PK\x03\x04' or is_zip_url

            if not is_zip:

                try:

                    return content.decode("utf-8-sig")

                except:

                    return content.decode("utf-8", errors="ignore")



            print("检测到 ZIP 压缩包，尝试解析……")

            csv_bytes = StockService._read_csv_from_zip(content, pwd_list)

            if csv_bytes is None:

                return None



            for enc in ("utf-8-sig", "utf-8", "gbk", "latin-1"):

                try:

                    return csv_bytes.decode(enc)

                except UnicodeDecodeError:

                    continue



            return csv_bytes.decode("utf-8", errors="ignore")



        except Exception as e:

            print(f"下载或解压异常：{e}")

            return None



    @staticmethod

    def _read_csv_from_zip(raw_zip_bytes: bytes, pwd_list: Iterable[str]) -> Optional[bytes]:

        # 1) zipfile

        try:

            with zipfile.ZipFile(io.BytesIO(raw_zip_bytes)) as zf:

                name = StockService._first_csv_name(zf.namelist())

                if not name: return None



                try:

                    with zf.open(name, "r") as f:

                        return f.read()

                except RuntimeError:

                    pass



                for p in pwd_list:

                    try:

                        with zf.open(name, "r", pwd=p.encode("utf-8")) as f:

                            return f.read()

                    except Exception:

                        continue

        except Exception:

            pass



        # 2) pyzipper (如果不装 pyzipper 这一段会跳过)

        try:

            import pyzipper

            with pyzipper.AESZipFile(io.BytesIO(raw_zip_bytes)) as zf:

                name = StockService._first_csv_name(zf.namelist())

                if not name: return None

                for p in pwd_list:

                    try:

                        zf.pwd = p.encode("utf-8")

                        with zf.open(name, "r") as f:

                            return f.read()

                    except:

                        continue

        except ImportError:

            print("提示: 未安装 pyzipper，无法处理 AES 加密 ZIP")

        except Exception:

            pass



        return None



    @staticmethod

    def _first_csv_name(names):

        for n in names:

            if n.lower().endswith(".csv"): return n

        return None



    @staticmethod

    def download_xlsx(url: str) -> io.BytesIO:

        session = requests.Session()

        session.trust_env = False  # ✅ 禁止系统代理（解决你日志里的 localhost warning）



        headers = {

            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",

            "Accept": "*/*",

            "Connection": "keep-alive",

        }



        print("开始下载:", url)

        resp = session.get(

            url,

            headers=headers,

            timeout=(10, 60),

            verify=True,  # ✅ 建议保持 True（S3 正常证书，不需要 verify=False）

            proxies={"http": None, "https": None},  # ✅ 再保险：强制不走代理

        )

        resp.raise_for_status()



        content_type = resp.headers.get("Content-Type", "")

        content = resp.content



        # ✅ xlsx 本质是 zip，文件头必须是 PK

        is_zip = content[:2] == b"PK"



        # ✅ 如果不是 zip，大概率就是下载到了 HTML 错误页（403/503/代理页）

        if not is_zip:

            preview = content[:300]

            try:

                preview = preview.decode("utf-8", errors="ignore")

            except:

                preview = str(preview)



            raise Exception(

                f"下载到的不是xlsx(不是zip)。Content-Type={content_type}\n"

                f"前300字节预览：\n{preview}"

            )



        print("✅ 下载成功：", len(content), "bytes", "Content-Type:", content_type)

        return io.BytesIO(content)



    # -------------------------------------------------

    #  数据处理逻辑 (调用 DBManager)

    # -------------------------------------------------



    @staticmethod

    def process_costway_data(csv_text: str) -> Tuple[bool, str]:

        """处理 Costway CSV -> newestdropship"""

        if not csv_text or not csv_text.strip():

            return False, "CSV 内容为空"



        try:

            df = pd.read_csv(StringIO(csv_text))

        except Exception as e:

            return False, f"请求发生异常: {e}", []



        required_cols = {'SKU', 'Price', 'Stock'}

        if required_cols - set(df.columns):

            return False, "缺少必要列"



        df = df[['SKU', 'Price', 'Stock']].copy()

        df['SKU'] = df['SKU'].fillna('').astype(str).str.strip()

        df['Price'] = pd.to_numeric(df['Price'], errors='coerce').fillna(999).astype(float)

        df['Stock'] = pd.to_numeric(df['Stock'], errors='coerce').fillna(999).astype(int)

        df = df[df['SKU'] != '']



        now = datetime.now()

        data = [(row['SKU'], float(row['Price']), int(row['Stock']), now) for _, row in df.iterrows()]



        try:

            # ★★★ 调用 DBManager 更新 Costway ★★★

            DBManager.update_costway_stock(data)

        except Exception as e:

            return False, f"请求发生异常: {e}", []



        return True, f"成功更新 {len(data)} 条"



    @staticmethod

    def process_songmics_data(csv_text: str) -> Tuple[bool, str]:

        """Process Songmics CSV -> newestdropship_songmics (SKU/Price/Stock)"""

        if not csv_text or not csv_text.strip():

            return False, "CSV content is empty"



        try:

            df = pd.read_csv(StringIO(csv_text))

        except Exception as e:

            return False, f"请求发生异常: {e}", []



        df.columns = [str(c).strip() for c in df.columns]



        if {"SKU", "Price", "Stock"}.issubset(df.columns):

            df = df[["SKU", "Price", "Stock"]].copy()

        elif {"SKU Price", "Stock"}.issubset(df.columns):

            sku_price = df["SKU Price"].astype(str).str.strip()

            price = sku_price.str.extract(r"(-?\d+(?:\.\d+)?)\s*$")[0]

            sku = sku_price.str.replace(r"(-?\d+(?:\.\d+)?)\s*$", "", regex=True).str.strip()

            df = pd.DataFrame({"SKU": sku, "Price": price, "Stock": df["Stock"]})

        else:

            return False, f"Missing required columns: SKU/Price/Stock or SKU Price/Stock. Columns: {df.columns.tolist()}"



        df["SKU"] = df["SKU"].fillna("").astype(str).str.strip()

        df["Price"] = pd.to_numeric(df["Price"], errors="coerce").fillna(0.0).astype(float)

        df["Stock"] = pd.to_numeric(df["Stock"], errors="coerce").fillna(0).astype(int)

        df = df[df["SKU"] != ""]



        now = datetime.now()

        data = [(row["SKU"], float(row["Price"]), int(row["Stock"]), now) for _, row in df.iterrows()]



        try:

            DBManager.update_songmics_stock(data)

        except Exception as e:

            return False, f"请求发生异常: {e}", []



        return True, f"Updated {len(data)} rows"



    @staticmethod

    def process_vevor_data(file_content) -> Tuple[bool, str]:

        """

        极速版：openpyxl read_only 流式读取，只取 3 列

        输出 data_tuples -> DBManager.rewrite_vevor_stock

        """

        try:

            if isinstance(file_content, str):

                # 你这里几乎用不到（保留兼容）

                import pandas as pd, io

                df = pd.read_csv(io.StringIO(file_content))

                df.columns = df.columns.str.strip()



                # 映射

                if 'after coupon price' in df.columns:

                    df['Price'] = df['after coupon price']

                if 'Inventory quantity' in df.columns:

                    df.rename(columns={'Inventory quantity': 'Stock'}, inplace=True)



                df = df[['SKU', 'Price', 'Stock']].copy()

                df['Price'] = df['Price'].astype(str).str.replace('USD', '', case=False).str.strip()

                df['Price'] = pd.to_numeric(df['Price'], errors='coerce').fillna(0.0)

                df['Stock'] = pd.to_numeric(df['Stock'], errors='coerce').fillna(0).astype(int)

                df['SKU'] = df['SKU'].astype(str).str.strip()

                df = df[(df['SKU'] != '') & (df['SKU'] != 'nan')]

                df.drop_duplicates(subset=['SKU'], keep='first', inplace=True)



                data_tuples = [(str(r.SKU), float(r.Price), int(r.Stock)) for r in df.itertuples(index=False)]

                DBManager.rewrite_vevor_stock(data_tuples)

                return True, f"成功重写 {len(data_tuples)} 条"



            # ===== xlsx 流式读取（重点）=====

            print("==[2] VEVOR openpyxl read_only loading...==")

            wb = load_workbook(file_content, read_only=True, data_only=True)

            ws = wb.active  # 默认第一张表



            rows = ws.iter_rows(values_only=True)

            header = next(rows, None)

            if not header:

                return False, "VEVOR Excel 表头为空"



            # 表头 -> 列索引（统一小写+去空格）

            def norm(x):

                return str(x).strip().lower() if x is not None else ""



            header_map = {norm(h): i for i, h in enumerate(header)}



            sku_idx = header_map.get("sku")

            stock_idx = header_map.get("inventory quantity")

            price_idx = header_map.get("after coupon price")  # 你要用这个



            if sku_idx is None or stock_idx is None or price_idx is None:

                return False, f"缺少必要列，当前表头: {list(header_map.keys())}"



            data_tuples = []

            seen = set()



            # 进度打印（每 5000 行打一条）

            n = 0

            for r in rows:

                n += 1

                if n % 5000 == 0:

                    print(f"==[2] VEVOR reading... rows={n}, valid={len(data_tuples)} ==")



                sku = r[sku_idx]

                if not sku:

                    continue

                sku = str(sku).strip()

                if not sku or sku.lower() == "nan":

                    continue

                if sku in seen:

                    continue

                seen.add(sku)



                stock = r[stock_idx]

                price = r[price_idx]



                # 清洗 price：可能是数字，也可能是 '335.90 USD'

                if isinstance(price, str):

                    price = price.replace("USD", "").strip()

                try:

                    price = float(price) if price is not None and price != "" else 0.0

                except:

                    price = 0.0



                try:

                    stock = int(stock) if stock is not None and stock != "" else 0

                except:

                    stock = 0



                data_tuples.append((sku, price, stock))



            if not data_tuples:

                return False, "无有效数据"



            print(f"==[2] VEVOR read done. rows={n}, valid={len(data_tuples)} ==")



            # 分批写库（你 rewrite_vevor_stock 已经支持 batch_size）

            DBManager.rewrite_vevor_stock(data_tuples)

            return True, f"成功重写 {len(data_tuples)} 条"



        except Exception as e:

            import traceback

            traceback.print_exc()

            return False, f"Vevor处理异常: {str(e)}"



    @staticmethod

    def sync_giga_stock():

        """

        同步 GIGA 库存主逻辑

        1. 读库获取 SKU

        2. 调 API

        3. 写库

        """

        try:

            # 1. 获取本地所有 SKU

            all_skus = DBManager.get_all_giga_skus()

            if not all_skus:

                return False, "GIGA 表为空，无法同步"



            print(f"GIGA: 待同步 SKU 数量: {len(all_skus)}")



            # 2. 分批调用 API (每次 200 个)

            batch_size = 200

            update_data = []  # 准备写入数据库的数据 [(price, stock, sku)]



            for i in range(0, len(all_skus), batch_size):

                batch = all_skus[i:i + batch_size]



                # A. 查库存

                inv_data = StockService._giga_post(

                    "/b2b-overseas-api/v1/buyer/inventory/quantity/v2",

                    {"skus": batch}

                )



                # 如果 inv_data 是 None 或者不是 dict 或者 success=False，就跳过这一批，不更新数据库

                if not isinstance(inv_data, dict):

                    print("GIGA inventory return None or invalid:", inv_data)

                    continue



                # inv_data 是 dict，但 success 失败也跳过

                if not inv_data.get("success"):

                    print("GIGA inventory success false:", inv_data)

                    continue



                inv_map = {}

                for item in inv_data.get("data", []):

                    if not isinstance(item, dict):

                        continue



                    sku = item.get("sku")

                    if not sku:

                        continue



                    seller_info = item.get("sellerInventoryInfo") or {}

                    if not isinstance(seller_info, dict):

                        seller_info = {}



                    inv_map[sku] = seller_info.get("sellerAvailableInventory", None)



                time.sleep(0.05)



                # B. 查价格

                price_data = StockService._giga_post(

                    "/b2b-overseas-api/v1/buyer/product/price/v1",

                    {"skus": batch}

                )



                if not isinstance(price_data, dict):

                    print("GIGA price return None or invalid:", price_data)

                    continue



                if not price_data.get("success"):

                    print("GIGA price success false:", price_data)

                    continue



                price_map = {}

                for item in price_data.get("data", []):

                    price_map[item["sku"]] = item.get("price", None)



                time.sleep(0.05)



                # C. 整合数据

                for sku in batch:

                    price = price_map.get(sku, None)  # 没有就 None

                    stock = inv_map.get(sku, None)  # 没有就 None



                    # 如果两个都是 None，说明这个SKU完全没拿到任何信息，就跳过

                    if price is None and stock is None:

                        continue



                    update_data.append((price, stock, sku))



            # 3. 批量更新数据库

            if update_data:

                DBManager.update_giga_stock(update_data)

                return True, f"GIGA: 成功更新 {len(update_data)} 条数据"

            else:

                return False, "GIGA: API 未返回有效数据"



        except Exception as e:

            import traceback

            traceback.print_exc()

            return False, f"GIGA 同步异常: {str(e)}"



    # --- GIGA 签名与请求工具 ---

    @staticmethod

    def _giga_post(path, body):

        try:

            timestamp = str(int(time.time() * 1000))

            nonce = "".join(random.choices(string.ascii_letters + string.digits, k=10))



            # 签名

            msg = f"{StockService.GIGA_CLIENT_ID}&{path}&{timestamp}&{nonce}"

            key = f"{StockService.GIGA_CLIENT_ID}&{StockService.GIGA_CLIENT_SECRET}&{nonce}"

            digest = hmac.new(key.encode("utf-8"), msg.encode("utf-8"), hashlib.sha256).digest()

            sign = base64.b64encode(digest.hex().encode("utf-8")).decode("utf-8")



            headers = {

                "Content-Type": "application/json",

                "client-id": StockService.GIGA_CLIENT_ID,

                "timestamp": timestamp,

                "nonce": nonce,

                "sign": sign,

            }



            url = StockService.GIGA_BASE_URL + path

            resp = requests.post(url, headers=headers, json=body, timeout=30)

            return resp.json()

        except Exception as e:

            print(f"GIGA API Error: {e}")

            return None







    @staticmethod
    def load_hd_config(base_dir: str) -> dict:
        config_path = os.path.join(base_dir, "instance", "hd_config.json")
        print(f"[HD_CONFIG][READ] {datetime.now().isoformat(timespec='seconds')} -> {config_path}")
        default_rules = {
            "haoya": {"threshold": 50, "qty_high": 20, "qty_low": 0},
            "sishun": {"threshold": 50, "qty_high": 20, "qty_low": 0},
            "dajian": {"threshold": 50, "qty_high": 20, "qty_low": 0},
            "songmics": {"threshold": 50, "qty_high": 20, "qty_low": 0},
        }
        default_cfg = {
            "excluded": [],
            "threshold": 50,
            "qty_high": 20,
            "qty_low": 0,
            "supplier_rules": default_rules,
            "costway_zip_password": "",
        }
        if os.path.exists(config_path):
            try:
                import json
                with open(config_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    default_cfg.update(
                        {
                            "excluded": data.get("excluded", []),
                            "threshold": data.get("threshold", 50),
                            "qty_high": data.get("qty_high", 20),
                            "qty_low": data.get("qty_low", 0),
                            "supplier_rules": data.get("supplier_rules", default_rules),
                            "costway_zip_password": data.get("costway_zip_password", ""),
                        }
                    )
            except Exception:
                pass
        return default_cfg

    @staticmethod
    def save_hd_config(base_dir: str, excluded_list, threshold=50, qty_high=20, qty_low=0, supplier_rules=None, costway_zip_password=None):
        config_path = os.path.join(base_dir, "instance", "hd_config.json")
        os.makedirs(os.path.dirname(config_path), exist_ok=True)
        rules = supplier_rules or {}
        # preserve existing password if not explicitly provided
        if costway_zip_password is None:
            existing = StockService.load_hd_config(base_dir)
            costway_zip_password = existing.get("costway_zip_password", "")
        data = {
            "excluded": list(excluded_list),
            "threshold": int(threshold),
            "qty_high": int(qty_high),
            "qty_low": int(qty_low),
            "supplier_rules": rules,
            "costway_zip_password": costway_zip_password,
        }
        import json
        print(f"[HD_CONFIG][WRITE] {datetime.now().isoformat(timespec='seconds')} -> {config_path}")
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    @staticmethod
    def _read_hd_api_token(base_dir: str) -> str:
        token = os.environ.get("TEAPPLIX_API_TOKEN", "")
        token_path = os.path.join(base_dir, "instance", "hd_api_token.txt")
        if os.path.exists(token_path):
            try:
                with open(token_path, "r", encoding="utf-8") as f:
                    file_token = f.read().strip()
                if file_token:
                    token = file_token
            except Exception:
                pass
        return token

    @staticmethod
    def _apply_supplier_rule(stock_val, supplier_key, cfg, default_threshold, default_high, default_low):
        if not supplier_key:
            return 0, "no_supplier -> 0"
        rules = {}
        if isinstance(cfg, dict):
            rules = cfg.get("supplier_rules", {}) or {}
        rule = rules.get(supplier_key, {}) if supplier_key else {}
        use_threshold = rule.get("threshold", default_threshold)
        use_high = rule.get("qty_high", default_high)
        use_low = rule.get("qty_low", default_low)
        if stock_val > use_threshold:
            return use_high, f">{use_threshold} -> {use_high}"
        return use_low, f"<={use_threshold} -> {use_low}"

    @staticmethod
    def sync_hd_inventory(base_dir: str, excluded_list, threshold=50, qty_high=20, qty_low=0):
        token = StockService._read_hd_api_token(base_dir)
        if not token:
            return False, "Missing API_TOKEN. Set it in instance/hd_api_token.txt", []

        headers = {
            "APIToken": token,
            "Content-Type": "application/json",
            "Accept": "application/json",
        }

        params = {"WarehouseId": StockService.HD_WAREHOUSE_ID}
        try:
            get_resp = requests.get(StockService.HD_BASE_URL, headers=headers, params=params, timeout=30)
            if get_resp.status_code != 200:
                return False, f"Fetch failed: {get_resp.text}", []

            items_data = get_resp.json().get("ProductQuantities", [])
            if not items_data:
                return False, "No products found in warehouse", []
        except Exception as e:
            return False, f"Request error: {e}", []

        update_list = []
        log_rows = []
        today = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

        excluded_set = set(excluded_list or [])
        cfg = StockService.load_hd_config(base_dir)
        supplier_rules = cfg.get("supplier_rules", {}) if isinstance(cfg, dict) else {}

        for row in items_data:
            item_name = row.get("ItemName")
            if not item_name:
                continue

            if item_name in excluded_set:
                db_qty = 0
                final_qty = 0
                reason = "excluded"
                supplier_key = None
            else:
                db_qty, supplier_key = DBManager.get_supplier_max_stock_with_source(item_name)
                rule = supplier_rules.get(supplier_key, {}) if supplier_key else {}
                use_threshold = rule.get("threshold", threshold)
                use_high = rule.get("qty_high", qty_high)
                use_low = rule.get("qty_low", qty_low)
                if db_qty > use_threshold:
                    final_qty = use_high
                    reason = f">{use_threshold} -> {use_high}"
                else:
                    final_qty = use_low
                    reason = f"<={use_threshold} -> {use_low}"

            update_list.append({
                "PostDate": today,
                "PostType": StockService.HD_POST_TYPE,
                "WarehouseId": StockService.HD_WAREHOUSE_ID,
                "ItemName": item_name,
                "Quantity": final_qty,
            })

            log_rows.append({
                "sku": item_name,
                "db_qty": db_qty,
                "final_qty": final_qty,
                "excluded": item_name in excluded_set,
                "reason": reason,
                "supplier": supplier_key or "",
            })

        if not update_list:
            return False, "No items to update", log_rows

        payload = {
            "Quantities": update_list,
            "Cleanup": False,
            "ProductCrossReference": "reject",
        }

        try:
            post_resp = requests.post(StockService.HD_BASE_URL, json=payload, headers=headers, timeout=60)
            if post_resp.status_code == 200:
                return True, f"HD sync success. Updated {len(update_list)} items", log_rows
            return False, f"Update failed: {post_resp.text} (tried {len(update_list)} items)", log_rows
        except Exception as e:
            return False, f"Update error: {e} (tried {len(update_list)} items)", log_rows

    def process_bestbuy_stock(file_path, output_path, base_dir=None):

        """

        处理 Bestbuy 库存逻辑 (保留原格式版)

        使用 openpyxl 直接修改单元格，不破坏表头颜色和布局

        """

        try:

            # 1. 获取库存字典

            stock_map = DBManager.get_shop_stock_max_map()
            cfg = StockService.load_hd_config(base_dir or os.getcwd())



            # 2. 使用 openpyxl 加载 Excel (这样能保留样式)

            wb = load_workbook(file_path)

            ws = wb.active  # 获取第一个工作表



            # 3. 寻找 'sku' 和 'quantity' 所在的列号

            sku_col_idx = None

            qty_col_idx = None



            # 遍历第一行表头，找到对应的列 (从1开始数)

            for cell in ws[1]:

                if not cell.value:

                    continue

                header = str(cell.value).strip().lower()



                if header == 'sku':

                    sku_col_idx = cell.column

                elif header == 'quantity':

                    qty_col_idx = cell.column



            if not sku_col_idx or not qty_col_idx:

                return False, f"表头错误：未找到 'sku' 或 'quantity' 列。请检查模板。"



            # 4. 遍历每一行数据 (从第2行开始)

            # ws.iter_rows 效率比逐个 cell 访问要高

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):

                # 获取 SKU 单元格对象

                sku_cell = row[sku_col_idx - 1]  # 索引从0开始，列号从1开始，所以减1

                qty_cell = row[qty_col_idx - 1]



                # 获取 SKU 的值

                shop_sku = str(sku_cell.value).strip() if sku_cell.value else ""



                if not shop_sku:

                    continue



                # 5. 查库存并应用规则

                stock_info = stock_map.get(shop_sku)
                if stock_info is None:
                    real_stock, supplier_key = 0, None
                else:
                    real_stock, supplier_key = stock_info

                new_qty, _reason = StockService._apply_supplier_rule(
                    real_stock,
                    supplier_key,
                    cfg,
                    default_threshold=50,
                    default_high=20,
                    default_low=0,
                )



                # 6. 直接修改 quantity 单元格的值

                qty_cell.value = new_qty



            # 7. 保存文件 (另存为 output_path)

            wb.save(output_path)



            return True, "Bestbuy 库存更新完成 (格式已保留)"



        except Exception as e:

            import traceback

            traceback.print_exc()

            return False, str(e)



    @staticmethod

    def process_macy_stock(file_path, output_path, base_dir=None):

        """

        处理 Macy 库存逻辑 (保留原格式版)

        使用 openpyxl 直接修改单元格，不破坏表头颜色和布局

        """

        try:

            # 1. 获取库存字典

            stock_map = DBManager.get_shop_stock_max_map()
            cfg = StockService.load_hd_config(base_dir or os.getcwd())



            # 2. 使用 openpyxl 加载 Excel (这样能保留样式)

            wb = load_workbook(file_path)



            # ✅ 最小化更改：优先使用指定工作表名(若存在)，否则用默认 active

            ws = wb["offers-import"] if "offers-import" in wb.sheetnames else wb.active



            # 3. 寻找 'sku' 和 'quantity' 所在的列号

            sku_col_idx = None

            qty_col_idx = None



            # 遍历第一行表头，找到对应的列 (从1开始数)

            for cell in ws[1]:

                if not cell.value:

                    continue

                header = str(cell.value).strip().lower()



                if header == 'sku':

                    sku_col_idx = cell.column

                elif header == 'quantity':

                    qty_col_idx = cell.column



            if not sku_col_idx or not qty_col_idx:

                return False, "表头错误：未找到 'sku' 或 'quantity' 列。请检查模板。"



            # 4. 遍历每一行数据 (从第2行开始)

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):

                sku_cell = row[sku_col_idx - 1]

                qty_cell = row[qty_col_idx - 1]



                shop_sku = str(sku_cell.value).strip() if sku_cell.value else ""

                if not shop_sku:

                    continue



                # ✅ 最小化更改：查库存时做一次兜底（避免库存表 SKU 大小写/空格差异）

                stock_info = stock_map.get(shop_sku)
                if stock_info is None:
                    stock_info = stock_map.get(shop_sku.upper())

                if stock_info is None:
                    real_stock, supplier_key = 0, None
                else:
                    real_stock, supplier_key = stock_info

                new_qty, _reason = StockService._apply_supplier_rule(
                    real_stock,
                    supplier_key,
                    cfg,
                    default_threshold=50,
                    default_high=20,
                    default_low=0,
                )



                # 6. 直接修改 quantity 单元格的值

                qty_cell.value = new_qty



            # 7. 保存文件 (另存为 output_path)

            wb.save(output_path)

            wb.close()



            return True, "Macy 库存更新完成 (格式已保留)"



        except Exception as e:

            import traceback

            traceback.print_exc()

            return False, str(e)

    @staticmethod
    def process_lowes_stock(file_path, output_path, base_dir=None):
        """处理 Lowes 库存逻辑，规则与 Macy 相同"""
        try:
            stock_map = DBManager.get_shop_stock_max_map()
            cfg = StockService.load_hd_config(base_dir or os.getcwd())

            wb = load_workbook(file_path)
            ws = wb["offers-import"] if "offers-import" in wb.sheetnames else wb.active

            sku_col_idx = None
            qty_col_idx = None
            for cell in ws[1]:
                if not cell.value:
                    continue
                header = str(cell.value).strip().lower()
                if header == 'sku':
                    sku_col_idx = cell.column
                elif header == 'quantity':
                    qty_col_idx = cell.column

            if not sku_col_idx or not qty_col_idx:
                return False, "表头错误：未找到 'sku' 或 'quantity' 列。请检查模板。"

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                sku_cell = row[sku_col_idx - 1]
                qty_cell = row[qty_col_idx - 1]
                shop_sku = str(sku_cell.value).strip() if sku_cell.value else ""
                if not shop_sku:
                    continue

                stock_info = stock_map.get(shop_sku)
                if stock_info is None:
                    stock_info = stock_map.get(shop_sku.upper())
                if stock_info is None:
                    real_stock, supplier_key = 0, None
                else:
                    real_stock, supplier_key = stock_info

                new_qty, _reason = StockService._apply_supplier_rule(
                    real_stock, supplier_key, cfg,
                    default_threshold=50, default_high=20, default_low=0,
                )
                qty_cell.value = new_qty

            wb.save(output_path)
            wb.close()
            return True, "Lowes 库存更新完成 (格式已保留)"
        except Exception as e:
            import traceback
            traceback.print_exc()
            return False, str(e)

